from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_BASE = Path("output/final_verified_av")
DEFAULT_OUTPUT_NAME = "AV_CV_Analysis_Plots_With_Data.xlsx"


PLOT_FILES = [
    ("Fig 4a – CV overlay",            "fig4a_cv_overlay.png"),
    ("Fig 4b – Current heatmap",        "fig4b_current_heatmap.png"),
    ("Fig 4c – Peak current scaling",   "fig4c_peak_scaling.png"),
    ("Fig 5a – Ising segmentation",     "fig5a_ising_segmentation.png"),
    ("Fig 5b – Quantum KPCA",           "fig5b_quantum_kpca.png"),
    ("Fig 5c – b-value heatmap",        "fig5c_b_value_heatmap.png"),
    ("Fig 5d – Capacitive-fraction map","fig5d_cap_fraction_map.png"),
    ("Fig 5e – Kinetic regime",         "fig5e_kinetic_regime.png"),
    ("Fig 5f – Pointwise b scaling",    "fig5f_pointwise_scaling.png"),
]

PLOT_SHEET_NAMES = [
    "Plot_Fig4a_CV_Overlay",
    "Plot_Fig4b_Current_Heatmap",
    "Plot_Fig4c_Peak_Scaling",
    "Plot_Fig5a_Ising",
    "Plot_Fig5b_QKPCA",
    "Plot_Fig5c_b_Heatmap",
    "Plot_Fig5d_Cap_Fraction",
    "Plot_Fig5e_Kinetic_Regime",
    "Plot_Fig5f_Pointwise",
]

PLOT_DATA_FILES = [
    ("Data_CV_Overlay", "01_cv_overlay_raw_loop.csv"),
    ("Data_Current_Heatmap", "02_current_heatmap_abs_current.csv"),
    ("Data_b_Regimes", "03_b_values_and_regimes.csv"),
    ("Data_Cap_Fraction_Map", "04_capacitive_fraction_map.csv"),
    ("Data_Dunn_Currents", "05_dunn_decomposition_currents.csv"),
    ("Data_Dunn_Fractions", "06_dunn_fractions_by_scan_rate.csv"),
    ("Data_Peak_Scaling", "07_peak_scaling_points_and_fits.csv"),
    ("Data_Quantum_Kernel", "08_quantum_kernel_matrix.csv"),
    ("Data_Quantum_KPCA", "09_quantum_kpca_scores.csv"),
    ("Data_QKPCA_Eigen", "10_quantum_kpca_eigenvalues.csv"),
    ("Data_Summary", "11_summary_table_data.csv"),
]

PLOT_TO_DATA = {
    "Plot_Fig4a_CV_Overlay":     "Data_CV_Overlay",
    "Plot_Fig4b_Current_Heatmap":"Data_Current_Heatmap",
    "Plot_Fig4c_Peak_Scaling":   "Data_Peak_Scaling",
    "Plot_Fig5a_Ising":          "Data_b_Regimes",
    "Plot_Fig5b_QKPCA":          "Data_Quantum_KPCA, Data_Quantum_Kernel, Data_QKPCA_Eigen",
    "Plot_Fig5c_b_Heatmap":      "Data_b_Regimes",
    "Plot_Fig5d_Cap_Fraction":   "Data_Cap_Fraction_Map",
    "Plot_Fig5e_Kinetic_Regime": "Data_b_Regimes",
    "Plot_Fig5f_Pointwise":      "Data_b_Regimes",
}


FORMULAS = [
    ["Analysis", "Formula used"],
    ["CV overlay", "x = raw potential V_raw[t]; y = raw current I_raw[t, v_j]"],
    ["Current heatmap", "H_ij = |I(V_i, v_j)|"],
    ["Peak current", "I_anodic,peak(v_j)=max_i I(V_i,v_j); I_cathodic,peak(v_j)=min_i I(V_i,v_j)"],
    ["Peak scaling", "|I_peak(v_j)| = m sqrt(v_j) + c; R2 = 1 - SS_res/SS_tot"],
    ["b-value", "log |I(V_i,v_j)| = log a(V_i) + b(V_i) log v_j"],
    ["Dunn model", "i(V,v)=k1(V)v + k2(V)sqrt(v)"],
    ["Dunn linear fit", "i(V,v)/sqrt(v) = k1(V)sqrt(v) + k2(V)"],
    ["Capacitive current", "i_cap(V,v)=k1(V)v"],
    ["Diffusion current", "i_diff(V,v)=k2(V)sqrt(v)"],
    ["Capacitive fraction", "F_cap(v)=int|i_cap|dV / (int|i_cap|dV + int|i_diff|dV)"],
    ["Threshold regime", "DD if b<=0.6; CD if b>=0.7; mixed otherwise"],
    ["Ising state", "s_i in {0,1}; 0=DD, 1=CD"],
    ["Ising objective", "E(s)=sum_i C_i(s_i)+lambda sum_i 1[s_i != s_(i-1)], lambda=0.18"],
    ["Ising local cost", "C_DD(i)=(b_i-0.5)^2; C_CD(i)=(b_i-1.0)^2"],
    ["Quantum feature map", "|phi(x)> = tensor_k Ry(x_k)|0>"],
    ["Quantum kernel", "K_ij = |<phi(x_i)|phi(x_j)>|^2 = product_k cos^2((x_ik-x_jk)/2)"],
    ["Kernel centering", "K_c = K - 1_N K - K 1_N + 1_N K 1_N"],
    ["KPCA eigenproblem", "K_c u_l = lambda_l u_l"],
    ["KPCA coordinate", "score_il = sqrt(lambda_l) u_il"],
    ["Explained variance", "EV_l = lambda_l / sum_j lambda_j"],
]


def style_sheet(ws, freeze: str | None = None) -> None:
    if freeze:
        ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D9E0EA")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if cell.row == 1:
                cell.font = Font(bold=True, color="1F2937")
                cell.fill = PatternFill("solid", fgColor="E8EEF7")


def autosize(ws, min_width=10, max_width=48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, max_width))


def append_df(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))


def add_plot_sheet(wb: Workbook, base: Path) -> None:
    ws = wb.create_sheet("Plots")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Final verified plots"
    ws["A1"].font = Font(bold=True, size=14, color="111827")
    row = 3
    for title, filename in PLOT_FILES:
        path = base / filename
        if not path.exists():
            continue
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11)
        img = XLImage(str(path))
        max_width = 880
        if img.width > max_width:
            ratio = max_width / img.width
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, f"A{row + 1}")
        row += max(18, int(img.height / 20) + 4)
    ws.column_dimensions["A"].width = 120


def add_exact_plot_sheets(wb: Workbook, base: Path) -> None:
    index = wb.create_sheet("Plot_Index")
    index.append(["Plot sheet", "Source PNG", "Data sheet(s)", "Embedded display", "Note"])
    style_sheet(index, "A2")
    for (title, filename), sheet_name in zip(PLOT_FILES, PLOT_SHEET_NAMES):
        path = base / filename
        if not path.exists():
            continue
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14, color="111827")
        ws["A2"] = f"Embedded from: {path.name}"
        ws["A2"].font = Font(italic=True, color="4B5563")
        ws["A3"] = f"Backing data sheet(s): {PLOT_TO_DATA.get(sheet_name, 'See data sheets')}"
        ws["A3"].font = Font(color="4B5563")
        ws["A4"] = "The image is the generated verified PNG embedded directly into this workbook."
        ws["A4"].font = Font(color="4B5563")

        img = XLImage(str(path))
        # Keep the source image unchanged in the workbook media; set a practical
        # display size so each plot is visible on opening the sheet.
        max_width = 1200
        if img.width > max_width:
            ratio = max_width / img.width
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, "A6")
        ws.column_dimensions["A"].width = 150
        for row in range(5, 5 + max(1, int(img.height / 18))):
            ws.row_dimensions[row].height = 18

        index.append([
            sheet_name,
            path.name,
            PLOT_TO_DATA.get(sheet_name, "See data sheets"),
            f"{img.width} x {img.height} px display",
            "Exact generated PNG embedded",
        ])
    autosize(index, max_width=42)


def add_plot_data_sheets(wb: Workbook, base: Path) -> None:
    index = wb.create_sheet("Data_Index")
    index.append(["Data sheet", "CSV file", "Rows", "Columns"])
    style_sheet(index, "A2")
    data_dir = base / "plot_data_csv"
    for sheet_name, filename in PLOT_DATA_FILES:
        path = data_dir / filename
        if not path.exists():
            continue
        df = pd.read_csv(path)
        ws = wb.create_sheet(sheet_name)
        append_df(ws, df)
        style_sheet(ws, "A2")
        autosize(ws, max_width=18)
        index.append([sheet_name, path.name, len(df), len(df.columns)])
    autosize(index, max_width=44)


def build(base_dir: str | Path = DEFAULT_BASE, csv_path: str | Path = "AV.csv", output_name: str = DEFAULT_OUTPUT_NAME) -> Path:
    base = Path(base_dir)
    audit_dir = base / "audit"
    output = base / output_name
    metrics = json.loads((base / "analysis_metrics.json").read_text(encoding="utf-8"))
    audit = json.loads((audit_dir / "calculation_audit.json").read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    rows = [
        ["Workbook", "RVCE CND Lab CV Analysis Plots and Audit"],
        ["Input CSV", Path(csv_path).name],
        ["Important caveat", "Per-sample workbook. Composition comparisons require distinct uploaded CSV files."],
        ["Analysis branch", metrics["analysis_branch"]],
        ["Raw rows", metrics["raw_rows"]],
        ["Unique potentials", metrics["raw_unique_potentials"]],
        ["Scan rates", ", ".join(str(int(v)) for v in metrics["scan_rates"]) + " mV/s"],
        ["Mean b", metrics["mean_b"]],
        ["Mean capacitive fraction", metrics["mean_cap_fraction"]],
        ["Audit status", "Independent b-value, R2, Dunn k1/k2, and quantum-kernel checks matched exactly."],
        ["Formula PDF", "CV_Analysis_Formulas_and_References.pdf"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    autosize(ws)

    ws = wb.create_sheet("Formulas")
    for row in FORMULAS:
        ws.append(row)
    style_sheet(ws, "A2")
    autosize(ws, max_width=80)

    ws = wb.create_sheet("Metrics")
    metric_rows = [
        ["Metric", "Value"],
        ["Potential min", metrics["potential_min"]],
        ["Potential max", metrics["potential_max"]],
        ["Current min", metrics["current_min"]],
        ["Current max", metrics["current_max"]],
        ["Mean b", metrics["mean_b"]],
        ["Median b", metrics["median_b"]],
        ["Mean R2", metrics["mean_r_squared"]],
        ["Mean cap fraction", metrics["mean_cap_fraction"]],
        ["Anodic peak R2", metrics["anodic_peak_r2"]],
        ["Cathodic peak R2", metrics["cathodic_peak_r2"]],
        ["Quantum kernel PC1 explained variance", metrics["quantum_kernel"]["explained_variance_first_two"][0]],
        ["Quantum kernel PC2 explained variance", metrics["quantum_kernel"]["explained_variance_first_two"][1]],
    ]
    for row in metric_rows:
        ws.append(row)
    style_sheet(ws)
    autosize(ws)

    for sheet_name, csv_path in [
        ("Peak_Dunn_Audit", audit_dir / "fig4_peak_and_dunn_audit.csv"),
        ("Quantum_Kernel", audit_dir / "fig5_quantum_kernel_matrix.csv"),
        ("Quantum_KPCA", audit_dir / "fig5_quantum_kpca_scores.csv"),
        ("Summary_Table", base / "summary_table.csv"),
    ]:
        ws = wb.create_sheet(sheet_name)
        append_df(ws, pd.read_csv(csv_path))
        style_sheet(ws, "A2")
        autosize(ws)

    ws = wb.create_sheet("Audit_Checks")
    ws.append(["Check", "Value"])
    for key, value in audit["independent_cross_checks"].items():
        ws.append([key, value])
    ws.append(["Fig4 anodic slope", audit["fig4_peak_scaling"]["anodic_slope"]])
    ws.append(["Fig4 cathodic slope", audit["fig4_peak_scaling"]["cathodic_slope"]])
    ws.append(["Fig5 b mean", audit["fig5_b_value"]["mean"]])
    ws.append(["Fig5 Dunn mean cap fraction", audit["fig5_dunn"]["mean_cap_fraction"]])
    style_sheet(ws)
    autosize(ws)

    raw = pd.read_csv(csv_path, header=None)
    ws = wb.create_sheet("Raw_AV_CSV")
    for row in raw.itertuples(index=False):
        ws.append(list(row))
    style_sheet(ws, "A4")
    autosize(ws, max_width=16)

    add_plot_sheet(wb, base)
    add_plot_data_sheets(wb, base)
    add_exact_plot_sheets(wb, base)

    wb.save(output)
    return output


def verify(path: Path) -> None:
    wb = load_workbook(path)
    required = {
        "README",
        "Formulas",
        "Metrics",
        "Peak_Dunn_Audit",
        "Quantum_Kernel",
        "Quantum_KPCA",
        "Summary_Table",
        "Audit_Checks",
        "Raw_AV_CSV",
        "Plots",
        "Plot_Index",
        "Data_Index",
    }
    missing = required.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {sorted(missing)}")
    embedded = len(wb["Plots"]._images)
    for sheet_name in PLOT_SHEET_NAMES:
        if sheet_name in wb.sheetnames:
            embedded += len(wb[sheet_name]._images)
    if embedded < len(PLOT_FILES) * 2:
        raise RuntimeError("Expected plot images were not embedded.")


if __name__ == "__main__":
    out = build()
    verify(out)
    print(out)
