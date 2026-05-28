from __future__ import annotations

import json
import math
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.optimize import least_squares, curve_fit
from scipy.signal import find_peaks, peak_widths, savgol_filter
from scipy.stats import linregress

from api_integrations import materials_project_lookup, nvidia_nim_commentary


# ─── 15 Publication-grade style presets ──────────────────────────────────────
STYLE_PRESETS: dict[str, dict] = {
    # ── Journal styles ────────────────────────────────────────────────────────
    "nature": {
        "font": "serif", "font_list": ["Times New Roman", "Times", "DejaVu Serif"],
        "cmap": "viridis", "grid": 0.20, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 9, "tick_dir": "in", "lw": 0.8,
        "palette": ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2", "#7f7f7f"],
    },
    "science": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "plasma", "grid": 0.18, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 9, "tick_dir": "out", "lw": 0.9,
        "palette": ["#003087", "#CC0000", "#007A33", "#FF6600", "#7B2FBE", "#00A0B0"],
    },
    "acs": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "turbo", "grid": 0.15, "face": "white",
        "fs": 7.5, "label_fs": 8.5, "title_fs": 8.5, "tick_dir": "in", "lw": 0.75,
        "palette": ["#0072B2", "#E69F00", "#009E73", "#CC79A7", "#56B4E9", "#D55E00", "#000000"],
    },
    "ieee": {
        "font": "serif", "font_list": ["Times New Roman", "Times", "DejaVu Serif"],
        "cmap": "Greys", "grid": 0.22, "face": "white",
        "fs": 8, "label_fs": 8, "title_fs": 8, "tick_dir": "in", "lw": 0.7,
        "palette": ["#000000", "#555555", "#AAAAAA", "#000000", "#777777"],
    },
    "elsevier": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "coolwarm", "grid": 0.20, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 9, "tick_dir": "out", "lw": 0.9,
        "palette": ["#1A4E99", "#E8472A", "#2E8B57", "#FF8C00", "#9B2FA5", "#00868A"],
    },
    "rsc": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "RdYlBu", "grid": 0.18, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 9, "tick_dir": "in", "lw": 0.85,
        "palette": ["#003366", "#CC0000", "#006633", "#FF6600", "#660099", "#006699"],
    },
    # ── General styles ───────────────────────────────────────────────────────
    "reference": {
        "font": "serif", "font_list": ["Times New Roman", "Times", "DejaVu Serif"],
        "cmap": "viridis", "grid": 0.24, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 10, "tick_dir": "in", "lw": 0.8,
        "palette": ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b"],
    },
    "minimal": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "magma", "grid": 0.10, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 9, "tick_dir": "out", "lw": 1.0,
        "palette": ["#2d2d2d", "#666666", "#999999", "#cccccc", "#444444", "#888888"],
    },
    "modern": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "viridis", "grid": 0.18, "face": "white",
        "fs": 9, "label_fs": 10, "title_fs": 10, "tick_dir": "out", "lw": 1.2,
        "palette": ["#6366F1", "#EC4899", "#14B8A6", "#F59E0B", "#EF4444", "#8B5CF6"],
    },
    "grayscale": {
        "font": "serif", "font_list": ["Times New Roman", "Times", "DejaVu Serif"],
        "cmap": "Greys", "grid": 0.22, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 9, "tick_dir": "in", "lw": 0.8,
        "palette": ["#000000", "#444444", "#888888", "#BBBBBB", "#333333", "#777777"],
    },
    "colorblind": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "cividis", "grid": 0.22, "face": "white",
        "fs": 8, "label_fs": 9, "title_fs": 9, "tick_dir": "in", "lw": 1.0,
        "palette": ["#0072B2", "#E69F00", "#009E73", "#CC79A7", "#56B4E9", "#D55E00"],
    },
    "dark": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "plasma", "grid": 0.18, "face": "#111827",
        "fs": 9, "label_fs": 10, "title_fs": 10, "tick_dir": "out", "lw": 1.1,
        "palette": ["#60A5FA", "#F472B6", "#34D399", "#FBBF24", "#A78BFA", "#FB923C"],
    },
    "high_contrast": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "hot", "grid": 0.30, "face": "white",
        "fs": 9, "label_fs": 10, "title_fs": 10, "tick_dir": "in", "lw": 1.4,
        "palette": ["#000000", "#FF0000", "#0000FF", "#008000", "#FF8800", "#800080"],
    },
    "presentation": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "turbo", "grid": 0.15, "face": "white",
        "fs": 11, "label_fs": 13, "title_fs": 13, "tick_dir": "out", "lw": 1.6,
        "palette": ["#2563EB", "#DC2626", "#16A34A", "#D97706", "#7C3AED", "#DB2777"],
    },
    "seaborn_white": {
        "font": "Arial", "font_list": ["Arial", "Helvetica", "DejaVu Sans"],
        "cmap": "tab10", "grid": 0.35, "face": "#f8f8f8",
        "fs": 9, "label_fs": 10, "title_fs": 10, "tick_dir": "out", "lw": 1.0,
        "palette": ["#4C72B0", "#DD8452", "#55A868", "#C44E52", "#8172B3", "#937860"],
    },
}

FORMULA_ROWS = [
    ["Technique", "Analysis", "Formula"],
    ["EIS", "Complex impedance", "Z(f) = Z'(f) + j Z''(f)"],
    ["EIS", "Magnitude", "|Z| = sqrt(Z'^2 + Z''^2)"],
    ["EIS", "Phase", "phi = atan2(Z'', Z') * 180/pi"],
    ["EIS", "Solution resistance", "Rs ~= Z'(f_max)"],
    ["EIS", "Charge-transfer resistance", "Rct ~= max(Z') - min(Z')"],
    ["EIS", "Randles RC fit", "Z = Rs + Rct / (1 + j omega Rct Cdl)"],
    ["EIS", "Characteristic frequency", "f0 = 1 / (2 pi Rct Cdl)"],
    ["EIS", "Complex capacitance C'", "C'(w) = -Z''(w) / (w |Z(w)|^2)"],
    ["EIS", "Complex capacitance C''", "C''(w) = Z'(w) / (w |Z(w)|^2)"],
    ["EIS", "Loss tangent", "tan(delta) = Z' / |Z''|"],
    ["DPV", "Baseline corrected", "Icorr(V) = I(V) - baseline(V)"],
    ["DPV", "Peak current", "Ip = max(Icorr)"],
    ["DPV", "Calibration", "Ip = m C + b"],
    ["DPV", "LOD", "LOD = 3 sigma_blank / |m|"],
    ["DPV", "LOQ", "LOQ = 10 sigma_blank / |m|"],
    ["GCD", "Specific capacitance", "Csp = I Delta t / (m Delta V)"],
    ["GCD", "IR drop", "Delta V_IR = V_before - V_after drop"],
    ["GCD", "Energy density", "E = 0.5 Csp Delta V^2 / 3.6"],
    ["GCD", "Power density", "P = E / Delta t * 3600"],
    ["GCD", "Coulombic efficiency", "eta = t_discharge / t_charge * 100"],
    ["Raman", "D/G ratio", "R = I_D / I_G"],
    ["Raman", "Crystallite size", "La = (2.4e-10) * lambda_L^4 * (I_G/I_D)"],
    ["Raman", "Lorentzian", "f(x) = A * (gamma/2)^2 / ((x-x0)^2 + (gamma/2)^2)"],
]


@dataclass
class GenericResult:
    technique: str
    sample_id: str
    source_name: str
    output_dir: Path
    metrics: dict
    plots: list[str]
    workbook: str


def apply_style(style: str) -> dict:
    preset = STYLE_PRESETS.get(style, STYLE_PRESETS["reference"])
    # Suppress font-not-found warnings for journal styles that request Arial/Times
    import warnings, logging
    logging.getLogger("matplotlib.font_manager").setLevel(logging.ERROR)
    warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")
    dark = preset["face"] not in {"white", "#f8f8f8"}
    fg = "white" if dark else "#111827"
    grid_c = "#374151" if dark else "#e5e7eb"
    matplotlib.rcParams.update(
        {
            "font.family": preset["font"],
            "font.serif": preset.get("font_list", ["DejaVu Serif"]),
            "font.sans-serif": preset.get("font_list", ["DejaVu Sans"]),
            "mathtext.fontset": "stix" if preset["font"] == "serif" else "dejavusans",
            "font.size": preset["fs"],
            "axes.labelsize": preset["label_fs"],
            "axes.titlesize": preset["title_fs"],
            "legend.fontsize": preset["fs"] - 0.5,
            "xtick.labelsize": preset["fs"] - 0.5,
            "ytick.labelsize": preset["fs"] - 0.5,
            "axes.linewidth": preset["lw"],
            "lines.linewidth": preset["lw"] + 0.2,
            "xtick.direction": preset["tick_dir"],
            "ytick.direction": preset["tick_dir"],
            "xtick.major.width": preset["lw"],
            "ytick.major.width": preset["lw"],
            "figure.facecolor": preset["face"],
            "savefig.facecolor": preset["face"],
            "axes.facecolor": preset["face"],
            "axes.edgecolor": fg,
            "axes.labelcolor": fg,
            "xtick.color": fg,
            "ytick.color": fg,
            "text.color": fg,
            "grid.color": grid_c,
            "pdf.fonttype": 42,
            "svg.fonttype": "none",
        }
    )
    return preset


def read_any(path: str | Path) -> pd.DataFrame:
    """Read CSV/XLSX/TXT/extensionless files into a raw DataFrame.

    Multer uploads files without extensions, so we must sniff by magic bytes
    when the suffix is missing or unrecognised.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    # Known extension → fast path
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, header=None)
    if suffix in {".txt", ".dat"}:
        return pd.read_csv(path, sep=None, engine="python", header=None, encoding_errors="replace")
    if suffix in {".csv"}:
        return pd.read_csv(path, header=None, encoding_errors="replace")

    # No extension (multer upload) or unknown → sniff by magic bytes
    with open(path, "rb") as fh:
        magic = fh.read(4)
    if magic[:2] == b"PK":  # ZIP-based: xlsx, xlsm, ods etc.
        return pd.read_excel(path, header=None)

    # Treat as text (CSV/TSV/space-separated)
    return pd.read_csv(path, sep=None, engine="python", header=None, encoding_errors="replace")


def _numeric_rows(df: pd.DataFrame, min_count: int = 2) -> pd.DataFrame:
    numeric = df.apply(pd.to_numeric, errors="coerce")
    return numeric[numeric.notna().sum(axis=1) >= min_count].dropna(axis=1, how="all")


def detect_technique(path: str | Path, requested: str = "auto") -> str:
    if requested and requested.lower() not in {"auto", ""}:
        return requested.lower()
    name = Path(path).name.lower()
    if "eis" in name or "impedance" in name or "nyquist" in name or "bode" in name:
        return "eis"
    if "dpv" in name or "differential" in name or "pulse" in name:
        return "dpv"
    if "gcd" in name or "charge" in name or "discharge" in name or "galvano" in name:
        return "gcd"
    if "raman" in name:
        return "raman"
    try:
        from loader import load_cv_dataset
        cv_data = load_cv_dataset(str(path))
        if len(cv_data.scan_rates) >= 2 and cv_data.currents_raw.shape[1] >= 2:
            return "cv"
    except Exception:
        pass
    try:
        df = read_any(path)
    except Exception:
        return "cv"
    text = " ".join(str(x).lower() for x in df.head(30).to_numpy().ravel())
    if "freq" in text and ("z'" in text or "phase" in text or "impedance" in text or "z_real" in text):
        return "eis"
    if "raman" in text or "shift" in text or ("intensity" in text and "cm" in text):
        return "raman"
    if "potential" in text and "current" in text and df.shape[1] >= 4:
        return "dpv"
    nums = _numeric_rows(df, min_count=2)
    if nums.shape[1] >= 4:
        return "dpv"
    return "gcd" if nums.shape[1] == 2 and "time" in text else "cv"


def _savefig(fig: plt.Figure, path: Path, dpi: int) -> str:
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    return path.name


def _write_json(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


def _zip_dir(source_dir: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in source_dir.rglob("*"):
            if path.is_file() and path != zip_path:
                zf.write(path, path.relative_to(source_dir))


def _colors_from_preset(preset: dict, n: int) -> list[str]:
    pal = preset.get("palette", ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd"])
    if n <= len(pal):
        return pal[:n]
    cmap = plt.get_cmap(preset.get("cmap", "viridis"))
    return [matplotlib.colors.to_hex(cmap(i / max(n - 1, 1))) for i in range(n)]


# ─── EIS Analysis ─────────────────────────────────────────────────────────────

def parse_eis(path: str | Path) -> pd.DataFrame:
    df = read_any(path)
    numeric = _numeric_rows(df, min_count=3)
    if numeric.shape[1] < 3 or len(numeric) < 5:
        raise ValueError("Could not find EIS columns (frequency, Z real, Z imaginary).")
    numeric = numeric.iloc[:, :5].copy()
    columns = ["frequency_hz", "z_real_ohm", "z_imag_ohm", "z_abs_ohm", "phase_deg"][: numeric.shape[1]]
    numeric.columns = columns
    numeric = numeric.dropna(subset=["frequency_hz", "z_real_ohm", "z_imag_ohm"])
    return numeric.sort_values("frequency_hz", ascending=False).reset_index(drop=True)


def _randles_model(params: np.ndarray, freq: np.ndarray) -> np.ndarray:
    rs, rct, cdl = params
    omega = 2 * np.pi * freq
    return rs + rct / (1 + 1j * omega * rct * cdl)


def analyze_eis(path: Path, output_dir: Path, dpi: int, style: str) -> tuple[dict, list[str], dict[str, pd.DataFrame]]:
    preset = apply_style(style)
    data = parse_eis(path)
    freq = data["frequency_hz"].to_numpy(float)
    zr = data["z_real_ohm"].to_numpy(float)
    zi = data["z_imag_ohm"].to_numpy(float)
    z_complex = zr + 1j * zi

    # ── Randles RC fit ────────────────────────────────────────────────────────
    rs0 = float(zr[np.argmax(freq)])
    rct0 = float(max(np.nanmax(zr) - np.nanmin(zr), 1e-9))
    cdl0 = 1e-6
    fit_ok = False
    fit = {"Rs_ohm": rs0, "Rct_ohm": rct0, "Cdl_F": cdl0, "f0_hz": None, "rmse_ohm": None}
    zfit = _randles_model(np.array([rs0, rct0, cdl0]), freq)
    try:
        def residual(p):
            z = _randles_model(p, freq)
            return np.r_[np.real(z - z_complex), np.imag(z - z_complex)]
        res = least_squares(residual, x0=np.array([rs0, rct0, cdl0]), bounds=([0, 0, 1e-12], [np.inf, np.inf, 1.0]), max_nfev=20000)
        zfit = _randles_model(res.x, freq)
        rmse = float(np.sqrt(np.mean(np.abs(zfit - z_complex) ** 2)))
        rs, rct, cdl = [float(x) for x in res.x]
        fit = {"Rs_ohm": rs, "Rct_ohm": rct, "Cdl_F": cdl,
               "f0_hz": float(1 / (2 * np.pi * rct * cdl)) if rct > 0 and cdl > 0 else None, "rmse_ohm": rmse}
        fit_ok = True
    except Exception:
        pass

    # ── Derived quantities ────────────────────────────────────────────────────
    omega = 2 * np.pi * freq
    z_sq = zr**2 + zi**2
    z_sq = np.where(z_sq < 1e-30, 1e-30, z_sq)
    C_prime = -zi / (omega * z_sq)          # C'(ω) — real capacitance (stores energy)
    C_dbl   =  zr / (omega * z_sq)          # C''(ω) — imaginary capacitance (dissipates)
    Y_prime = zr / z_sq                      # G (conductance)
    Y_dbl   = -zi / z_sq                     # B (susceptance)
    loss_tangent = np.abs(zr / np.where(np.abs(zi) < 1e-30, 1e-30, zi))
    residual_real = np.real(z_complex - zfit)
    residual_imag = np.imag(z_complex - zfit)
    residual_pct = 100 * np.abs(z_complex - zfit) / np.where(np.abs(z_complex) < 1e-30, 1e-30, np.abs(z_complex))

    # ── Processed data CSV ────────────────────────────────────────────────────
    data_out = data.copy()
    data_out["minus_z_imag_ohm"] = -zi
    data_out["z_fit_real_ohm"] = np.real(zfit)
    data_out["z_fit_imag_ohm"] = np.imag(zfit)
    data_out["C_prime_F"] = C_prime
    data_out["C_dbl_prime_F"] = C_dbl
    data_out["Y_prime_S"] = Y_prime
    data_out["Y_dbl_S"] = Y_dbl
    data_out["loss_tangent"] = loss_tangent
    data_out["residual_pct"] = residual_pct
    data_out.to_csv(output_dir / "eis_processed_data.csv", index=False)

    plots = []
    accent = preset.get("palette", ["#2563EB"])[0]
    accent2 = preset.get("palette", ["#2563EB", "#DC2626"])[1] if len(preset.get("palette", [])) > 1 else "#DC2626"
    grid_a = preset["grid"]

    # ── 1. Nyquist ────────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(5.8, 4.8), constrained_layout=True)
    ax.plot(zr, -zi, "o-", ms=3.5, lw=1.1, color=accent, label="Measured")
    if fit_ok:
        ax.plot(np.real(zfit), -np.imag(zfit), "--", lw=1.2, color=accent2, label=f"Randles fit (RMSE={fit['rmse_ohm']:.3g} Ω)")
        ax.scatter([fit["Rs_ohm"]], [0], s=40, marker="^", color=accent2, zorder=5, label=f"Rs = {fit['Rs_ohm']:.3g} Ω")
    ax.set_xlabel("Z' (Ω)")
    ax.set_ylabel("-Z'' (Ω)")
    ax.set_title("Nyquist Plot")
    ax.grid(True, alpha=grid_a)
    ax.legend(fontsize=preset["fs"] - 1)
    ax.set_aspect("equal", adjustable="datalim")
    plots.append(_savefig(fig, output_dir / "eis_nyquist.png", dpi))

    # ── 2. Bode ───────────────────────────────────────────────────────────────
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(6.2, 5.5), sharex=True, constrained_layout=True)
    ax1.loglog(freq, np.abs(z_complex), "o-", ms=3.0, color=accent, lw=1.0)
    if fit_ok:
        ax1.loglog(freq, np.abs(zfit), "--", lw=1.0, color=accent2, alpha=0.8)
    ax1.set_ylabel("|Z| (Ω)")
    ax1.set_title("Bode Plot")
    ax1.grid(True, which="both", alpha=grid_a)
    ax2.semilogx(freq, np.degrees(np.angle(z_complex)), "o-", ms=3.0, color=accent, lw=1.0)
    if fit_ok:
        ax2.semilogx(freq, np.degrees(np.angle(zfit)), "--", lw=1.0, color=accent2, alpha=0.8)
    ax2.set_xlabel("Frequency (Hz)")
    ax2.set_ylabel("Phase angle (°)")
    ax2.grid(True, which="both", alpha=grid_a)
    plots.append(_savefig(fig, output_dir / "eis_bode.png", dpi))

    # ── 3. Z components vs frequency ──────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(6.2, 3.8), constrained_layout=True)
    ax.semilogx(freq, zr, "o-", ms=3.0, color=accent, lw=1.0, label="Z' (real)")
    ax.semilogx(freq, -zi, "s-", ms=3.0, color=accent2, lw=1.0, label="-Z'' (imag)")
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("Impedance (Ω)")
    ax.set_title("Impedance Components")
    ax.grid(True, which="both", alpha=grid_a)
    ax.legend()
    plots.append(_savefig(fig, output_dir / "eis_components.png", dpi))

    # ── 4. Complex capacitance C' and C'' ────────────────────────────────────
    valid = np.isfinite(C_prime) & np.isfinite(C_dbl) & (freq > 0)
    if valid.sum() >= 4:
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(6.2, 5.5), sharex=True, constrained_layout=True)
        ax1.semilogx(freq[valid], C_prime[valid], "o-", ms=3.0, color=accent, lw=1.0)
        ax1.set_ylabel("C' (F)")
        ax1.set_title("Complex Capacitance")
        ax1.grid(True, which="both", alpha=grid_a)
        ax2.semilogx(freq[valid], C_dbl[valid], "s-", ms=3.0, color=accent2, lw=1.0)
        ax2.set_xlabel("Frequency (Hz)")
        ax2.set_ylabel("C'' (F)")
        ax2.grid(True, which="both", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "eis_capacitance.png", dpi))

        # ── 5. Capacitance Cole-Cole (C'' vs C') ─────────────────────────────
        fig, ax = plt.subplots(figsize=(5.5, 4.5), constrained_layout=True)
        sc = ax.scatter(C_prime[valid], C_dbl[valid], c=np.log10(np.maximum(freq[valid], 1e-10)),
                        cmap=preset["cmap"], s=18)
        plt.colorbar(sc, ax=ax, label="log₁₀ f (Hz)")
        ax.set_xlabel("C' (F)")
        ax.set_ylabel("C'' (F)")
        ax.set_title("Capacitance Cole-Cole")
        ax.grid(True, alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "eis_cole_cole.png", dpi))

    # ── 6. Admittance Y' and Y'' ─────────────────────────────────────────────
    valid_y = np.isfinite(Y_prime) & np.isfinite(Y_dbl)
    if valid_y.sum() >= 4:
        fig, ax = plt.subplots(figsize=(6.2, 3.8), constrained_layout=True)
        ax.loglog(freq[valid_y], np.abs(Y_prime[valid_y]), "o-", ms=3.0, color=accent, lw=1.0, label="G (conductance)")
        ax.loglog(freq[valid_y], np.abs(Y_dbl[valid_y]), "s-", ms=3.0, color=accent2, lw=1.0, label="|B| (susceptance)")
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel("Admittance (S)")
        ax.set_title("Complex Admittance")
        ax.grid(True, which="both", alpha=grid_a)
        ax.legend()
        plots.append(_savefig(fig, output_dir / "eis_admittance.png", dpi))

    # ── 7. Randles fit residuals ──────────────────────────────────────────────
    if fit_ok:
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(6.2, 5.0), sharex=True, constrained_layout=True)
        ax1.semilogx(freq, residual_real, "o-", ms=3.0, color=accent, lw=0.9)
        ax1.axhline(0, lw=0.7, ls="--", color="gray")
        ax1.set_ylabel("ΔZ' (Ω)")
        ax1.set_title("Randles Fit Residuals")
        ax1.grid(True, which="both", alpha=grid_a)
        ax2.semilogx(freq, residual_imag, "s-", ms=3.0, color=accent2, lw=0.9)
        ax2.axhline(0, lw=0.7, ls="--", color="gray")
        ax2.set_xlabel("Frequency (Hz)")
        ax2.set_ylabel("ΔZ'' (Ω)")
        ax2.grid(True, which="both", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "eis_fit_residuals.png", dpi))

    # ── 8. Loss tangent ───────────────────────────────────────────────────────
    valid_lt = np.isfinite(loss_tangent) & (loss_tangent < 1e6)
    if valid_lt.sum() >= 4:
        fig, ax = plt.subplots(figsize=(6.2, 3.8), constrained_layout=True)
        ax.semilogx(freq[valid_lt], loss_tangent[valid_lt], "o-", ms=3.0, color=accent, lw=1.0)
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel("tan δ = Z'/|Z''|")
        ax.set_title("Dielectric Loss Tangent")
        ax.grid(True, which="both", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "eis_loss_tangent.png", dpi))

    # ── 9. Relaxation frequency map ───────────────────────────────────────────
    if valid.sum() >= 4:
        max_c2_idx = int(np.nanargmax(np.abs(C_dbl[valid])))
        f_relax = float(freq[valid][max_c2_idx])
        fig, ax = plt.subplots(figsize=(6.2, 3.8), constrained_layout=True)
        ax.semilogx(freq[valid], np.abs(C_dbl[valid]), "o-", ms=3.0, color=accent, lw=1.0)
        ax.axvline(f_relax, ls="--", lw=0.9, color=accent2, label=f"f₀ = {f_relax:.3g} Hz")
        ax.set_xlabel("Frequency (Hz)")
        ax.set_ylabel("|C''| (F)")
        ax.set_title("Relaxation Frequency from C''")
        ax.legend()
        ax.grid(True, which="both", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "eis_relaxation.png", dpi))
    else:
        f_relax = None

    # ── 10. 4-panel publication summary ──────────────────────────────────────
    fig = plt.figure(figsize=(9.0, 7.0), constrained_layout=True)
    gs = gridspec.GridSpec(2, 2, figure=fig)
    ax_n = fig.add_subplot(gs[0, 0])
    ax_b = fig.add_subplot(gs[0, 1])
    ax_c = fig.add_subplot(gs[1, 0])
    ax_l = fig.add_subplot(gs[1, 1])
    # Nyquist
    ax_n.plot(zr, -zi, "o-", ms=2.5, lw=0.9, color=accent, label="Data")
    if fit_ok:
        ax_n.plot(np.real(zfit), -np.imag(zfit), "--", lw=0.9, color=accent2, label="Fit")
    ax_n.set_xlabel("Z' (Ω)")
    ax_n.set_ylabel("-Z'' (Ω)")
    ax_n.set_title("(a) Nyquist")
    ax_n.grid(True, alpha=grid_a)
    ax_n.legend(fontsize=preset["fs"] - 1.5)
    ax_n.set_aspect("equal", adjustable="datalim")
    # Bode
    ax_b.loglog(freq, np.abs(z_complex), "o-", ms=2.5, lw=0.9, color=accent)
    ax_b.set_xlabel("Frequency (Hz)")
    ax_b.set_ylabel("|Z| (Ω)")
    ax_b.set_title("(b) Bode |Z|")
    ax_b.grid(True, which="both", alpha=grid_a)
    # C' capacitance
    if valid.sum() >= 4:
        ax_c.semilogx(freq[valid], C_prime[valid], "o-", ms=2.5, lw=0.9, color=accent)
        ax_c.set_xlabel("Frequency (Hz)")
        ax_c.set_ylabel("C' (F)")
        ax_c.set_title("(c) Real Capacitance")
        ax_c.grid(True, which="both", alpha=grid_a)
    # Loss tangent
    if valid_lt.sum() >= 4:
        ax_l.semilogx(freq[valid_lt], loss_tangent[valid_lt], "o-", ms=2.5, lw=0.9, color=accent2)
        ax_l.set_xlabel("Frequency (Hz)")
        ax_l.set_ylabel("tan δ")
        ax_l.set_title("(d) Loss Tangent")
        ax_l.grid(True, which="both", alpha=grid_a)
    plots.append(_savefig(fig, output_dir / "eis_summary.png", dpi))

    metrics = {
        "technique": "EIS",
        "points": int(len(data)),
        "frequency_min_hz": float(np.nanmin(freq)),
        "frequency_max_hz": float(np.nanmax(freq)),
        "Rs_high_frequency_ohm": rs0,
        "Rct_diameter_estimate_ohm": rct0,
        "randles_fit": fit,
        "fit_ok": fit_ok,
        "relaxation_frequency_hz": f_relax,
    }
    return metrics, plots, {"EIS_Data": data_out}


# ─── DPV Analysis ─────────────────────────────────────────────────────────────

def parse_pairs(path: str | Path) -> list[tuple[str, pd.DataFrame]]:
    df = read_any(path)
    labels = []
    for col in range(0, df.shape[1] - 1, 2):
        label = None
        for row in range(min(4, len(df))):
            value = df.iloc[row, col + 1]
            if pd.notna(value):
                label = str(value).strip()
                break
        if not label or label.lower() in {"nan", ""}:
            label = f"series_{col//2 + 1}"
        x = pd.to_numeric(df.iloc[:, col], errors="coerce")
        y = pd.to_numeric(df.iloc[:, col + 1], errors="coerce")
        pair = pd.DataFrame({"x": x, "y": y}).dropna()
        if len(pair) >= 8:
            labels.append((label, pair.reset_index(drop=True)))
    if not labels:
        numeric = _numeric_rows(df, min_count=2)
        if numeric.shape[1] >= 2:
            labels.append((Path(path).stem, numeric.iloc[:, :2].dropna().rename(
                columns={numeric.columns[0]: "x", numeric.columns[1]: "y"})))
    return labels


def _baseline(y: np.ndarray) -> np.ndarray:
    window = min(len(y) // 2 * 2 - 1, 101)
    if window < 7:
        return np.full_like(y, np.nanmedian(y))
    return savgol_filter(y, window_length=window, polyorder=2)


def _label_to_concentration(label: str) -> float | None:
    try:
        return float(str(label).split()[0])
    except Exception:
        return None


def analyze_dpv(path: Path, output_dir: Path, dpi: int, style: str) -> tuple[dict, list[str], dict[str, pd.DataFrame]]:
    preset = apply_style(style)
    pairs = parse_pairs(path)
    if not pairs:
        raise ValueError("Could not find DPV potential/current pairs.")

    colors = _colors_from_preset(preset, len(pairs))
    processed_rows = []
    peaks = []
    plots = []
    grid_a = preset["grid"]

    # ── 1. Raw overlay ────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(6.5, 4.4), constrained_layout=True)
    for (label, pair), col in zip(pairs, colors):
        x = pair["x"].to_numpy(float)
        y = pair["y"].to_numpy(float)
        ax.plot(x, y, lw=1.0, label=label, color=col)
        base = _baseline(y)
        corr = y - base
        prominence = max(np.nanstd(corr) * 1.5, np.nanmax(np.abs(corr)) * 0.03, 1e-15)
        peak_idx, _ = find_peaks(corr, prominence=prominence)
        if len(peak_idx) == 0:
            peak_idx = np.array([int(np.nanargmax(corr))])
        best = int(peak_idx[np.nanargmax(corr[peak_idx])])
        widths = peak_widths(corr, [best], rel_height=0.5)
        dx = float(np.nanmedian(np.abs(np.diff(x)))) if len(x) > 1 else 1.0
        fwhm = float(widths[0][0] * dx)
        area = float(np.trapezoid(np.maximum(corr, 0), x))
        concentration = _label_to_concentration(label)
        peaks.append({
            "series": label, "concentration": concentration,
            "peak_potential_V": float(x[best]), "peak_current": float(y[best]),
            "baseline_corrected_peak": float(corr[best]),
            "area": area, "fwhm_V": fwhm,
        })
        for xi, yi, bi, ci in zip(x, y, base, corr):
            processed_rows.append({"series": label, "potential_V": xi, "current": yi, "baseline": bi, "current_corrected": ci})
    ax.set_xlabel("Potential (V)")
    ax.set_ylabel("Current")
    ax.set_title("DPV Raw Overlay")
    ax.grid(True, alpha=grid_a)
    ax.legend(ncol=2, fontsize=preset["fs"] - 1)
    plots.append(_savefig(fig, output_dir / "dpv_overlay.png", dpi))

    processed = pd.DataFrame(processed_rows)
    peak_df = pd.DataFrame(peaks)
    processed.to_csv(output_dir / "dpv_processed_data.csv", index=False)
    peak_df.to_csv(output_dir / "dpv_peak_table.csv", index=False)

    # ── 2. Baseline corrected + marked peaks ─────────────────────────────────
    fig, ax = plt.subplots(figsize=(6.5, 4.4), constrained_layout=True)
    for (label, group), col in zip(processed.groupby("series", sort=False), colors):
        ax.plot(group["potential_V"], group["current_corrected"], lw=1.0, label=label, color=col)
    ax.scatter(peak_df["peak_potential_V"], peak_df["baseline_corrected_peak"],
               s=30, color="#b91c1c", zorder=4, marker="v", label="Peaks")
    ax.set_xlabel("Potential (V)")
    ax.set_ylabel("Baseline-corrected current")
    ax.set_title("DPV Baseline-Corrected Peaks")
    ax.grid(True, alpha=grid_a)
    ax.legend(ncol=2, fontsize=preset["fs"] - 1)
    plots.append(_savefig(fig, output_dir / "dpv_baseline_corrected.png", dpi))

    # ── 3. Peak positions scatter ─────────────────────────────────────────────
    if len(peak_df) >= 1:
        fig, ax = plt.subplots(figsize=(6.0, 3.5), constrained_layout=True)
        bar_colors = colors[:len(peak_df)]
        bars = ax.bar(range(len(peak_df)), peak_df["baseline_corrected_peak"], color=bar_colors)
        ax.set_xticks(range(len(peak_df)))
        ax.set_xticklabels([str(s)[:12] for s in peak_df["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 1)
        ax.set_ylabel("Peak current (baseline-corrected)")
        ax.set_title("DPV Peak Currents by Series")
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "dpv_peak_currents.png", dpi))

    # ── 4. Peak areas bar chart ───────────────────────────────────────────────
    if len(peak_df) >= 1:
        fig, ax = plt.subplots(figsize=(6.0, 3.5), constrained_layout=True)
        ax.bar(range(len(peak_df)), peak_df["area"], color=colors[:len(peak_df)])
        ax.set_xticks(range(len(peak_df)))
        ax.set_xticklabels([str(s)[:12] for s in peak_df["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 1)
        ax.set_ylabel("Peak area (∫ Icorr dV)")
        ax.set_title("DPV Peak Areas by Series")
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "dpv_peak_areas.png", dpi))

    # ── 5. FWHM analysis ──────────────────────────────────────────────────────
    if len(peak_df) >= 1:
        fig, ax = plt.subplots(figsize=(6.0, 3.5), constrained_layout=True)
        ax.bar(range(len(peak_df)), peak_df["fwhm_V"] * 1000, color=colors[:len(peak_df)])
        ax.set_xticks(range(len(peak_df)))
        ax.set_xticklabels([str(s)[:12] for s in peak_df["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 1)
        ax.set_ylabel("FWHM (mV)")
        ax.set_title("DPV Peak FWHM by Series")
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "dpv_fwhm.png", dpi))

    # ── 6. Calibration curve ─────────────────────────────────────────────────
    calibration = pd.DataFrame()
    cal_metrics = None
    concentrations = peak_df["concentration"].dropna()
    if len(concentrations) >= 3:
        cal = peak_df.dropna(subset=["concentration"]).sort_values("concentration")
        x = cal["concentration"].to_numpy(float)
        y = cal["baseline_corrected_peak"].to_numpy(float)
        slope, intercept = np.polyfit(x, y, 1)
        yhat = slope * x + intercept
        r2 = 1 - np.sum((y - yhat) ** 2) / max(np.sum((y - np.mean(y)) ** 2), 1e-30)
        blank = peak_df[peak_df["series"].str.contains("buffer|blank", case=False, na=False)]["baseline_corrected_peak"]
        sigma_blank = float(np.std(blank, ddof=1)) if len(blank) >= 2 else float(np.std(y - yhat, ddof=1)) if len(y) >= 3 else None
        lod = float(3 * sigma_blank / abs(slope)) if sigma_blank and slope != 0 else None
        loq = float(10 * sigma_blank / abs(slope)) if sigma_blank and slope != 0 else None
        calibration = pd.DataFrame({"concentration": x, "peak_current_corrected": y, "fit": yhat})
        calibration.to_csv(output_dir / "dpv_calibration.csv", index=False)
        fig, ax = plt.subplots(figsize=(5.8, 4.2), constrained_layout=True)
        ax.plot(x, y, "o", color=colors[0], label="Peaks", ms=6)
        ax.plot(x, yhat, "-", color=colors[1] if len(colors) > 1 else "#dc2626", label=f"Linear fit R²={r2:.4f}")
        if lod is not None:
            ax.axvline(lod, ls=":", lw=0.8, color="gray", label=f"LOD={lod:.3g}")
        if loq is not None:
            ax.axvline(loq, ls="--", lw=0.8, color="gray", label=f"LOQ={loq:.3g}")
        ax.set_xlabel("Concentration")
        ax.set_ylabel("Baseline-corrected peak current")
        ax.set_title("DPV Calibration Curve")
        ax.grid(True, alpha=grid_a)
        ax.legend(fontsize=preset["fs"] - 1)
        plots.append(_savefig(fig, output_dir / "dpv_calibration.png", dpi))
        cal_metrics = {"slope": float(slope), "intercept": float(intercept), "r_squared": float(r2),
                       "sigma_blank_or_residual": sigma_blank, "LOD": lod, "LOQ": loq}

        # ── 7. LOD/LOQ visualization ──────────────────────────────────────────
        if lod is not None or loq is not None:
            x_ext = np.linspace(0, float(np.nanmax(x)) * 1.3, 200)
            fig, ax = plt.subplots(figsize=(5.8, 3.8), constrained_layout=True)
            ax.plot(x_ext, slope * x_ext + intercept, "-", color=colors[0], lw=1.2)
            ax.plot(x, y, "o", color=colors[0], ms=5, zorder=4)
            if lod is not None:
                ax.axvspan(0, lod, alpha=0.12, color="#ef4444", label=f"Below LOD ({lod:.3g})")
            if loq is not None:
                ax.axvspan(lod or 0, loq, alpha=0.08, color="#f59e0b", label=f"LOQ region ({loq:.3g})")
            ax.set_xlabel("Concentration")
            ax.set_ylabel("Peak current")
            ax.set_title("DPV Detection Limits")
            ax.legend(fontsize=preset["fs"] - 1)
            ax.grid(True, alpha=grid_a)
            plots.append(_savefig(fig, output_dir / "dpv_lod_loq.png", dpi))

    # ── 8. Summary panel ─────────────────────────────────────────────────────
    if len(pairs) >= 1:
        fig = plt.figure(figsize=(9.0, 6.5), constrained_layout=True)
        gs_s = gridspec.GridSpec(2, 2, figure=fig)
        ax_ov = fig.add_subplot(gs_s[0, 0])
        ax_bc = fig.add_subplot(gs_s[0, 1])
        ax_pk = fig.add_subplot(gs_s[1, 0])
        ax_ar = fig.add_subplot(gs_s[1, 1])
        for (label, group), col in zip(processed.groupby("series", sort=False), colors):
            ax_ov.plot(group["potential_V"], group["current"], lw=0.9, label=label[:10], color=col)
            ax_bc.plot(group["potential_V"], group["current_corrected"], lw=0.9, color=col)
        ax_ov.set_title("(a) Raw Overlay")
        ax_ov.set_xlabel("Potential (V)")
        ax_ov.set_ylabel("Current")
        ax_ov.grid(True, alpha=grid_a)
        ax_bc.set_title("(b) Baseline Corrected")
        ax_bc.set_xlabel("Potential (V)")
        ax_bc.set_ylabel("Corrected current")
        ax_bc.grid(True, alpha=grid_a)
        ax_pk.bar(range(len(peak_df)), peak_df["baseline_corrected_peak"], color=colors[:len(peak_df)])
        ax_pk.set_title("(c) Peak Currents")
        ax_pk.set_xticks(range(len(peak_df)))
        ax_pk.set_xticklabels([str(s)[:8] for s in peak_df["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 2)
        ax_pk.grid(True, axis="y", alpha=grid_a)
        ax_ar.bar(range(len(peak_df)), peak_df["area"], color=colors[:len(peak_df)])
        ax_ar.set_title("(d) Peak Areas")
        ax_ar.set_xticks(range(len(peak_df)))
        ax_ar.set_xticklabels([str(s)[:8] for s in peak_df["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 2)
        ax_ar.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "dpv_summary.png", dpi))

    metrics = {
        "technique": "DPV",
        "series_count": len(pairs),
        "peak_count": int(len(peak_df)),
        "calibration": cal_metrics,
    }
    tables = {"DPV_Processed": processed, "DPV_Peaks": peak_df}
    if not calibration.empty:
        tables["DPV_Calibration"] = calibration
    return metrics, plots, tables


# ─── GCD Analysis ─────────────────────────────────────────────────────────────

def _parse_current_density(label: str) -> float | None:
    """Try to extract a current density value (A/g) from a series label."""
    import re
    m = re.search(r"([\d.]+)\s*(a/g|ma/g|a g|ma g|a·g|ma·g)", label.lower())
    if m:
        val = float(m.group(1))
        if "ma" in m.group(2):
            val /= 1000
        return val
    m = re.search(r"([\d.]+)", label)
    return float(m.group(1)) if m else None


def analyze_gcd(path: Path, output_dir: Path, dpi: int, style: str) -> tuple[dict, list[str], dict[str, pd.DataFrame]]:
    preset = apply_style(style)
    pairs = parse_pairs(path)
    if not pairs:
        raise ValueError("Could not find GCD time/potential columns.")

    colors = _colors_from_preset(preset, len(pairs))
    rows = []
    summary_rows = []
    plots = []
    grid_a = preset["grid"]

    # ── 1. Charge-discharge profiles ─────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(6.4, 4.4), constrained_layout=True)
    for (label, pair), col in zip(pairs, colors):
        t = pair["x"].to_numpy(float)
        v = pair["y"].to_numpy(float)
        order = np.argsort(t)
        t, v = t[order], v[order]
        ax.plot(t, v, lw=1.0, label=label, color=col)
        dvdt = np.gradient(v, t) if len(t) > 2 else np.zeros_like(v)
        dt = float(np.nanmax(t) - np.nanmin(t))
        dv = float(np.nanmax(v) - np.nanmin(v))
        v_max = float(np.nanmax(v))
        v_min = float(np.nanmin(v))
        # IR drop: look for initial sharp drop at start of discharge
        discharge_mask = dvdt < np.nanmedian(dvdt)
        dis_v = v[discharge_mask]
        ir_drop = float(v_max - dis_v[0]) if len(dis_v) > 0 else 0.0
        # Charge / discharge time split
        t_half = float(np.nanmedian(t))
        t_charge = float(np.sum(np.diff(t[t <= t_half]))) if np.any(t <= t_half) else dt / 2
        t_discharge = float(np.sum(np.diff(t[t > t_half]))) if np.any(t > t_half) else dt / 2
        ce = float(t_discharge / t_charge * 100) if t_charge > 0 else None
        current_density = _parse_current_density(label)
        csp = float(current_density * t_discharge / max(dv, 1e-15)) if current_density and dv > 0 else None
        energy = float(0.5 * (csp or 0) * dv**2 / 3.6) if csp else None
        power = float(energy / max(t_discharge / 3600, 1e-15)) if energy and t_discharge > 0 else None
        summary_rows.append({
            "series": label, "duration_s": dt, "voltage_window_V": dv,
            "v_max_V": v_max, "v_min_V": v_min, "ir_drop_V": ir_drop,
            "t_charge_s": t_charge, "t_discharge_s": t_discharge,
            "coulombic_efficiency_pct": ce, "current_density_A_g": current_density,
            "specific_capacitance_F_g": csp, "energy_density_Wh_kg": energy, "power_density_W_kg": power,
        })
        for ti, vi, si in zip(t, v, dvdt):
            rows.append({"series": label, "time_s": ti, "potential_V": vi, "dVdt_V_s": si})
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Potential (V)")
    ax.set_title("Galvanostatic Charge-Discharge")
    ax.grid(True, alpha=grid_a)
    ax.legend(fontsize=preset["fs"] - 1, ncol=2)
    plots.append(_savefig(fig, output_dir / "gcd_profiles.png", dpi))

    data = pd.DataFrame(rows)
    summary = pd.DataFrame(summary_rows)
    data.to_csv(output_dir / "gcd_processed_data.csv", index=False)
    summary.to_csv(output_dir / "gcd_summary.csv", index=False)

    # ── 2. dV/dt slope ───────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(6.4, 4.0), constrained_layout=True)
    for (label, group), col in zip(data.groupby("series", sort=False), colors):
        t = group["time_s"].to_numpy(float)
        tn = (t - np.nanmin(t)) / max(np.nanmax(t) - np.nanmin(t), 1e-30)
        ax.plot(tn, group["dVdt_V_s"], lw=0.9, label=label[:12], color=col)
    ax.set_xlabel("Normalized time")
    ax.set_ylabel("dV/dt (V/s)")
    ax.set_title("GCD Differential Slope")
    ax.grid(True, alpha=grid_a)
    ax.legend(fontsize=preset["fs"] - 1, ncol=2)
    plots.append(_savefig(fig, output_dir / "gcd_dvdt.png", dpi))

    # ── 3. IR drop analysis ──────────────────────────────────────────────────
    ir_series = summary[summary["ir_drop_V"].notna() & (summary["ir_drop_V"] > 0)]
    if len(ir_series) >= 1:
        fig, ax = plt.subplots(figsize=(6.0, 3.5), constrained_layout=True)
        ax.bar(range(len(ir_series)), ir_series["ir_drop_V"] * 1000, color=colors[:len(ir_series)])
        ax.set_xticks(range(len(ir_series)))
        ax.set_xticklabels([str(s)[:12] for s in ir_series["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 1)
        ax.set_ylabel("IR drop (mV)")
        ax.set_title("GCD IR Drop by Series")
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "gcd_ir_drop.png", dpi))

    # ── 4. Specific capacitance bar chart ─────────────────────────────────────
    csp_series = summary[summary["specific_capacitance_F_g"].notna()]
    if len(csp_series) >= 1:
        fig, ax = plt.subplots(figsize=(6.0, 3.5), constrained_layout=True)
        ax.bar(range(len(csp_series)), csp_series["specific_capacitance_F_g"], color=colors[:len(csp_series)])
        ax.set_xticks(range(len(csp_series)))
        ax.set_xticklabels([str(s)[:12] for s in csp_series["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 1)
        ax.set_ylabel("Specific capacitance (F/g)")
        ax.set_title("GCD Specific Capacitance")
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "gcd_capacitance.png", dpi))

        # ── 5. Ragone plot ─────────────────────────────────────────────────────
        e_series = csp_series[csp_series["energy_density_Wh_kg"].notna() & csp_series["power_density_W_kg"].notna()]
        if len(e_series) >= 2:
            fig, ax = plt.subplots(figsize=(5.8, 4.4), constrained_layout=True)
            sc = ax.scatter(e_series["power_density_W_kg"], e_series["energy_density_Wh_kg"],
                            c=range(len(e_series)), cmap=preset["cmap"], s=50, zorder=4)
            for i, row in enumerate(e_series.itertuples()):
                ax.annotate(str(row.series)[:10], (row.power_density_W_kg, row.energy_density_Wh_kg),
                            fontsize=preset["fs"] - 2, ha="center", va="bottom")
            ax.set_xlabel("Power density (W/kg)")
            ax.set_ylabel("Energy density (Wh/kg)")
            ax.set_title("Ragone Plot")
            ax.grid(True, alpha=grid_a)
            plots.append(_savefig(fig, output_dir / "gcd_ragone.png", dpi))

    # ── 6. Rate capability ────────────────────────────────────────────────────
    rate_series = summary[summary["current_density_A_g"].notna() & summary["specific_capacitance_F_g"].notna()].sort_values("current_density_A_g")
    if len(rate_series) >= 2:
        fig, ax = plt.subplots(figsize=(5.8, 3.8), constrained_layout=True)
        ax.plot(rate_series["current_density_A_g"], rate_series["specific_capacitance_F_g"],
                "o-", ms=5, color=colors[0], lw=1.1)
        ax.set_xlabel("Current density (A/g)")
        ax.set_ylabel("Specific capacitance (F/g)")
        ax.set_title("Rate Capability")
        ax.grid(True, alpha=grid_a)
        # Capacitance retention
        c_first = float(rate_series["specific_capacitance_F_g"].iloc[0])
        c_last = float(rate_series["specific_capacitance_F_g"].iloc[-1])
        retention = 100 * c_last / max(c_first, 1e-15)
        ax.text(0.97, 0.97, f"Retention: {retention:.1f}%", transform=ax.transAxes,
                ha="right", va="top", fontsize=preset["fs"] - 1, bbox=dict(boxstyle="round", alpha=0.2))
        plots.append(_savefig(fig, output_dir / "gcd_rate_capability.png", dpi))

    # ── 7. Coulombic efficiency ───────────────────────────────────────────────
    ce_series = summary[summary["coulombic_efficiency_pct"].notna()]
    if len(ce_series) >= 1:
        fig, ax = plt.subplots(figsize=(6.0, 3.5), constrained_layout=True)
        ax.bar(range(len(ce_series)), ce_series["coulombic_efficiency_pct"], color=colors[:len(ce_series)])
        ax.axhline(100, ls="--", lw=0.8, color="gray", alpha=0.7)
        ax.set_xticks(range(len(ce_series)))
        ax.set_xticklabels([str(s)[:12] for s in ce_series["series"]], rotation=30, ha="right", fontsize=preset["fs"] - 1)
        ax.set_ylabel("Coulombic efficiency (%)")
        ax.set_ylim(0, max(110, float(ce_series["coulombic_efficiency_pct"].max()) * 1.1))
        ax.set_title("Coulombic Efficiency")
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "gcd_coulombic_eff.png", dpi))

    # ── 8. GCD summary panel ──────────────────────────────────────────────────
    fig = plt.figure(figsize=(9.0, 7.0), constrained_layout=True)
    gs_g = gridspec.GridSpec(2, 2, figure=fig)
    ax_pr = fig.add_subplot(gs_g[0, 0])
    ax_dv = fig.add_subplot(gs_g[0, 1])
    ax_ir = fig.add_subplot(gs_g[1, 0])
    ax_ce = fig.add_subplot(gs_g[1, 1])
    for (label, group), col in zip(data.groupby("series", sort=False), colors):
        ax_pr.plot(group["time_s"], group["potential_V"], lw=0.9, label=label[:10], color=col)
        t = group["time_s"].to_numpy(float)
        tn = (t - np.nanmin(t)) / max(np.nanmax(t) - np.nanmin(t), 1e-30)
        ax_dv.plot(tn, group["dVdt_V_s"], lw=0.9, color=col)
    ax_pr.set_title("(a) GCD Profiles")
    ax_pr.set_xlabel("Time (s)")
    ax_pr.set_ylabel("Potential (V)")
    ax_pr.grid(True, alpha=grid_a)
    ax_dv.set_title("(b) dV/dt")
    ax_dv.set_xlabel("Normalized time")
    ax_dv.set_ylabel("dV/dt (V/s)")
    ax_dv.grid(True, alpha=grid_a)
    if len(ir_series) >= 1:
        ax_ir.bar(range(len(ir_series)), ir_series["ir_drop_V"] * 1000, color=colors[:len(ir_series)])
        ax_ir.set_title("(c) IR Drop")
        ax_ir.set_ylabel("IR drop (mV)")
        ax_ir.grid(True, axis="y", alpha=grid_a)
    if len(ce_series) >= 1:
        ax_ce.bar(range(len(ce_series)), ce_series["coulombic_efficiency_pct"], color=colors[:len(ce_series)])
        ax_ce.axhline(100, ls="--", lw=0.8, color="gray", alpha=0.7)
        ax_ce.set_title("(d) Coulombic Eff.")
        ax_ce.set_ylabel("η (%)")
        ax_ce.grid(True, axis="y", alpha=grid_a)
    plots.append(_savefig(fig, output_dir / "gcd_summary.png", dpi))

    metrics = {
        "technique": "GCD",
        "series_count": len(pairs),
        "summary": summary.to_dict(orient="records"),
    }
    return metrics, plots, {"GCD_Data": data, "GCD_Summary": summary}


# ─── Raman Analysis ───────────────────────────────────────────────────────────

def _lorentzian(x: np.ndarray, amp: float, cen: float, wid: float) -> np.ndarray:
    return amp * (wid / 2) ** 2 / ((x - cen) ** 2 + (wid / 2) ** 2)


def _fit_raman_band(x: np.ndarray, y: np.ndarray, center_guess: float, window: float = 120.0) -> dict | None:
    mask = (x >= center_guess - window) & (x <= center_guess + window)
    if mask.sum() < 5:
        return None
    xf, yf = x[mask], y[mask]
    peak_amp = float(yf.max())
    try:
        popt, pcov = curve_fit(
            _lorentzian, xf, yf,
            p0=[peak_amp, center_guess, 30.0],
            bounds=([0, center_guess - window, 1.0], [peak_amp * 10, center_guess + window, window * 2]),
            maxfev=8000,
        )
        amp, cen, fwhm = popt
        perr = np.sqrt(np.diag(pcov))
        residuals = yf - _lorentzian(xf, *popt)
        ss_res = float(np.sum(residuals**2))
        ss_tot = float(np.sum((yf - yf.mean()) ** 2))
        r2 = 1 - ss_res / max(ss_tot, 1e-30)
        return {"amplitude": float(amp), "center_cm1": float(cen), "fwhm_cm1": float(abs(fwhm)), "r2": float(r2), "center_err": float(perr[1])}
    except Exception:
        return None


def _als_baseline(y: np.ndarray, lam: float = 1e5, p: float = 0.01, n_iter: int = 20) -> np.ndarray:
    """Asymmetric least squares baseline estimation."""
    from scipy.sparse import diags, eye as speye
    from scipy.sparse.linalg import spsolve
    m = len(y)
    D = diags([1, -2, 1], [0, 1, 2], shape=(m - 2, m)).toarray()
    H = lam * D.T @ D
    w = np.ones(m)
    z = y.copy()
    for _ in range(n_iter):
        W = diags(w, 0)
        try:
            from scipy.sparse import csr_matrix
            from scipy.sparse.linalg import spsolve as sp_spsolve
            z = sp_spsolve(csr_matrix(W + H), w * y)
        except Exception:
            z = np.linalg.solve(np.diag(w) + H, w * y)
        w = np.where(y > z, p, 1 - p)
    return z


def analyze_raman(path: Path, output_dir: Path, dpi: int, style: str) -> tuple[dict, list[str], dict[str, pd.DataFrame]]:
    preset = apply_style(style)
    df = read_any(path)
    numeric = _numeric_rows(df, min_count=2).iloc[:, :2].dropna()
    if numeric.shape[1] < 2:
        raise ValueError(f"Raman file must have at least 2 numeric columns (Raman shift, intensity). Found {numeric.shape[1]}.")
    numeric = numeric.copy()
    numeric.columns = ["raman_shift_cm_1", "intensity"]
    numeric = numeric.sort_values("raman_shift_cm_1").reset_index(drop=True)
    numeric.to_csv(output_dir / "raman_processed_data.csv", index=False)

    x = numeric["raman_shift_cm_1"].to_numpy(float)
    y = numeric["intensity"].to_numpy(float)
    plots = []
    grid_a = preset["grid"]
    accent = preset.get("palette", ["#2563EB"])[0]
    accent2 = preset.get("palette", ["#2563EB", "#DC2626"])[1] if len(preset.get("palette", [])) > 1 else "#DC2626"

    # ── Background subtraction ────────────────────────────────────────────────
    try:
        baseline = _als_baseline(y)
        y_sub = y - baseline
        y_sub = y_sub - y_sub.min()
    except Exception:
        window = min(len(y) // 4 * 2 - 1, 201)
        baseline = savgol_filter(y, window_length=max(window, 7), polyorder=2) if len(y) > 7 else np.zeros_like(y)
        y_sub = y - baseline
        y_sub = y_sub - y_sub.min()

    bg_df = pd.DataFrame({"raman_shift_cm_1": x, "intensity_raw": y, "baseline": baseline, "intensity_sub": y_sub})
    bg_df.to_csv(output_dir / "raman_background_subtracted.csv", index=False)

    # ── Peak finding on subtracted spectrum ───────────────────────────────────
    prom = max(np.std(y_sub), np.ptp(y_sub) * 0.04, 1.0)
    peaks_idx, _ = find_peaks(y_sub, prominence=prom, distance=max(1, len(y_sub) // 50))
    peak_df = pd.DataFrame({
        "raman_shift_cm_1": x[peaks_idx],
        "intensity_raw": y[peaks_idx],
        "intensity_sub": y_sub[peaks_idx],
    }).sort_values("intensity_sub", ascending=False).reset_index(drop=True)
    peak_df.to_csv(output_dir / "raman_peaks.csv", index=False)

    # Band assignments
    D_BAND = 1350.0
    G_BAND = 1580.0
    BAND_2D = 2700.0

    # ── 1. Full spectrum with peaks ───────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(7.0, 4.2), constrained_layout=True)
    ax.plot(x, y_sub, lw=1.0, color=accent)
    top_peaks = peak_df.head(12)
    ax.scatter(top_peaks["raman_shift_cm_1"], top_peaks["intensity_sub"], s=22, color=accent2, zorder=4)
    for _, row in top_peaks.head(8).iterrows():
        ax.annotate(f"{row['raman_shift_cm_1']:.0f}", (row["raman_shift_cm_1"], row["intensity_sub"]),
                    textcoords="offset points", xytext=(0, 5), ha="center", fontsize=preset["fs"] - 2)
    ax.set_xlabel("Raman shift (cm⁻¹)")
    ax.set_ylabel("Intensity (a.u.)")
    ax.set_title("Raman Spectrum")
    ax.grid(True, alpha=grid_a)
    plots.append(_savefig(fig, output_dir / "raman_spectrum.png", dpi))

    # ── 2. Background-subtracted vs raw comparison ────────────────────────────
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(7.0, 5.5), sharex=True, constrained_layout=True)
    ax1.plot(x, y, lw=0.9, color=accent, label="Raw")
    ax1.plot(x, baseline, lw=0.9, ls="--", color=accent2, alpha=0.8, label="Baseline")
    ax1.set_ylabel("Intensity (a.u.)")
    ax1.set_title("Background Subtraction")
    ax1.legend(fontsize=preset["fs"] - 1)
    ax1.grid(True, alpha=grid_a)
    ax2.plot(x, y_sub, lw=0.9, color=accent)
    ax2.set_xlabel("Raman shift (cm⁻¹)")
    ax2.set_ylabel("Corrected intensity (a.u.)")
    ax2.grid(True, alpha=grid_a)
    plots.append(_savefig(fig, output_dir / "raman_background_sub.png", dpi))

    # ── 3. D band zoom + Lorentzian fit ──────────────────────────────────────
    d_fit = _fit_raman_band(x, y_sub, D_BAND, window=150.0)
    if d_fit and np.any((x >= D_BAND - 200) & (x <= D_BAND + 200)):
        mask_d = (x >= D_BAND - 200) & (x <= D_BAND + 200)
        fig, ax = plt.subplots(figsize=(5.5, 3.8), constrained_layout=True)
        ax.plot(x[mask_d], y_sub[mask_d], "o-", ms=2.0, lw=0.8, color=accent, label="Data")
        x_fit = np.linspace(x[mask_d].min(), x[mask_d].max(), 300)
        ax.plot(x_fit, _lorentzian(x_fit, d_fit["amplitude"], d_fit["center_cm1"], d_fit["fwhm_cm1"]),
                "--", lw=1.2, color=accent2, label=f"Lorentzian\n{d_fit['center_cm1']:.1f} cm⁻¹, FWHM={d_fit['fwhm_cm1']:.1f}")
        ax.set_xlabel("Raman shift (cm⁻¹)")
        ax.set_ylabel("Intensity (a.u.)")
        ax.set_title("D Band (~1350 cm⁻¹)")
        ax.legend(fontsize=preset["fs"] - 1.5)
        ax.grid(True, alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "raman_d_band.png", dpi))

    # ── 4. G band zoom + Lorentzian fit ──────────────────────────────────────
    g_fit = _fit_raman_band(x, y_sub, G_BAND, window=120.0)
    if g_fit and np.any((x >= G_BAND - 150) & (x <= G_BAND + 150)):
        mask_g = (x >= G_BAND - 150) & (x <= G_BAND + 150)
        fig, ax = plt.subplots(figsize=(5.5, 3.8), constrained_layout=True)
        ax.plot(x[mask_g], y_sub[mask_g], "o-", ms=2.0, lw=0.8, color=accent, label="Data")
        x_fit = np.linspace(x[mask_g].min(), x[mask_g].max(), 300)
        ax.plot(x_fit, _lorentzian(x_fit, g_fit["amplitude"], g_fit["center_cm1"], g_fit["fwhm_cm1"]),
                "--", lw=1.2, color=accent2, label=f"Lorentzian\n{g_fit['center_cm1']:.1f} cm⁻¹, FWHM={g_fit['fwhm_cm1']:.1f}")
        ax.set_xlabel("Raman shift (cm⁻¹)")
        ax.set_ylabel("Intensity (a.u.)")
        ax.set_title("G Band (~1580 cm⁻¹)")
        ax.legend(fontsize=preset["fs"] - 1.5)
        ax.grid(True, alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "raman_g_band.png", dpi))

    # ── 5. 2D band ────────────────────────────────────────────────────────────
    band_2d_fit = _fit_raman_band(x, y_sub, BAND_2D, window=200.0)
    x_range_2d = np.any((x >= BAND_2D - 250) & (x <= BAND_2D + 250))
    if band_2d_fit and x_range_2d:
        mask_2d = (x >= BAND_2D - 250) & (x <= BAND_2D + 250)
        fig, ax = plt.subplots(figsize=(5.5, 3.8), constrained_layout=True)
        ax.plot(x[mask_2d], y_sub[mask_2d], "o-", ms=2.0, lw=0.8, color=accent, label="Data")
        x_fit = np.linspace(x[mask_2d].min(), x[mask_2d].max(), 300)
        ax.plot(x_fit, _lorentzian(x_fit, band_2d_fit["amplitude"], band_2d_fit["center_cm1"], band_2d_fit["fwhm_cm1"]),
                "--", lw=1.2, color=accent2, label=f"Lorentzian\n{band_2d_fit['center_cm1']:.1f} cm⁻¹")
        ax.set_xlabel("Raman shift (cm⁻¹)")
        ax.set_ylabel("Intensity (a.u.)")
        ax.set_title("2D Band (~2700 cm⁻¹)")
        ax.legend(fontsize=preset["fs"] - 1.5)
        ax.grid(True, alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "raman_2d_band.png", dpi))

    # ── 6. Second derivative for peak resolution ──────────────────────────────
    if len(y_sub) > 20:
        smooth_y = savgol_filter(y_sub, window_length=min(21, len(y_sub) // 5 * 2 + 1), polyorder=3)
        deriv2 = np.gradient(np.gradient(smooth_y, x), x)
        fig, ax = plt.subplots(figsize=(7.0, 3.5), constrained_layout=True)
        ax.plot(x, -deriv2, lw=0.9, color=accent)
        ax.axhline(0, lw=0.6, ls="--", color="gray", alpha=0.6)
        ax.set_xlabel("Raman shift (cm⁻¹)")
        ax.set_ylabel("-d²I/dx² (a.u.)")
        ax.set_title("Second Derivative (Peak Resolution Enhancement)")
        ax.grid(True, alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "raman_second_derivative.png", dpi))

    # ── 7. D/G ratio & defect density ────────────────────────────────────────
    I_D = d_fit["amplitude"] if d_fit else None
    I_G = g_fit["amplitude"] if g_fit else None
    dg_ratio = float(I_D / max(I_G, 1e-30)) if I_D and I_G else None
    LASER_NM = 532.0
    La_nm = float((2.4e-10) * (LASER_NM * 1e-9) ** 4 * (max(I_G, 1e-30) / max(I_D, 1e-30)) * 1e9) if dg_ratio else None

    if dg_ratio is not None:
        fig, ax = plt.subplots(figsize=(5.0, 4.2), constrained_layout=True)
        bars = ax.bar(["D band", "G band"], [I_D or 0, I_G or 0], color=[accent, accent2], width=0.5)
        for bar, val in zip(bars, [I_D or 0, I_G or 0]):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() * 1.01, f"{val:.3g}",
                    ha="center", va="bottom", fontsize=preset["fs"] - 1)
        ax.set_ylabel("Lorentzian amplitude (a.u.)")
        ax.set_title(f"D/G Intensity Ratio = {dg_ratio:.3f}" + (f"\nLa ≈ {La_nm:.1f} nm (λ={LASER_NM:.0f} nm)" if La_nm else ""))
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "raman_dg_ratio.png", dpi))

    # ── 8. Peak summary bar chart (top 10 peaks) ──────────────────────────────
    if len(peak_df) >= 2:
        top10 = peak_df.head(10)
        fig, ax = plt.subplots(figsize=(7.0, 3.8), constrained_layout=True)
        c_vals = _colors_from_preset(preset, len(top10))
        bars = ax.bar(range(len(top10)), top10["intensity_sub"], color=c_vals)
        ax.set_xticks(range(len(top10)))
        ax.set_xticklabels([f"{v:.0f}" for v in top10["raman_shift_cm_1"]], rotation=45, ha="right", fontsize=preset["fs"] - 1)
        ax.set_xlabel("Raman shift (cm⁻¹)")
        ax.set_ylabel("Intensity (a.u.)")
        ax.set_title("Top Raman Peaks")
        ax.grid(True, axis="y", alpha=grid_a)
        plots.append(_savefig(fig, output_dir / "raman_peak_summary.png", dpi))

    metrics = {
        "technique": "Raman",
        "points": int(len(numeric)),
        "peak_count": int(len(peak_df)),
        "top_peaks_cm1": peak_df["raman_shift_cm_1"].head(10).tolist(),
        "D_band_fit": d_fit,
        "G_band_fit": g_fit,
        "band_2D_fit": band_2d_fit,
        "DG_ratio": dg_ratio,
        "crystallite_size_La_nm": La_nm,
    }
    return metrics, plots, {"Raman_Data": numeric, "Raman_Background": bg_df, "Raman_Peaks": peak_df}


# ─── Excel workbook builder ───────────────────────────────────────────────────

def build_generic_workbook(output_dir: Path, source_name: str, metrics: dict, plots: list[str], tables: dict[str, pd.DataFrame]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    for row in [
        ["Workbook", "RVCE CND Lab Electrochemistry Analysis"],
        ["Source", source_name],
        ["Technique", metrics.get("technique")],
        ["Audit note", "Automatically generated — inspect raw and processed sheets before publication."],
    ]:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80

    ws = wb.create_sheet("Formulas")
    for row in FORMULA_ROWS:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8EEF7")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70

    ws = wb.create_sheet("Metrics")
    ws.append(["Metric", "Value"])
    for key, value in metrics.items():
        ws.append([key, json.dumps(value, default=str) if isinstance(value, (dict, list)) else value])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90

    ws = wb.create_sheet("Plots")
    row = 1
    for plot in plots:
        path = output_dir / plot
        if not path.exists():
            continue
        ws.cell(row=row, column=1, value=plot).font = Font(bold=True)
        img = XLImage(str(path))
        max_width = 900
        if img.width > max_width:
            ratio = max_width / img.width
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, f"A{row + 1}")
        row += max(18, int(img.height / 20) + 4)
    ws.column_dimensions["A"].width = 120

    for sheet_name, df in tables.items():
        safe = sheet_name[:31]
        ws = wb.create_sheet(safe)
        ws.append(list(df.columns))
        for row_data in df.itertuples(index=False):
            ws.append(list(row_data))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8EEF7")
        ws.freeze_panes = "A2"

    workbook = output_dir / "electrochemistry_analysis.xlsx"
    wb.save(workbook)
    return workbook.name


# ─── analyze_file ─────────────────────────────────────────────────────────────

def analyze_file(
    path: str | Path,
    output_dir: str | Path,
    sample_id: str | None = None,
    technique: str = "auto",
    dpi: int = 900,
    style: str = "reference",
    material_query: str | None = None,
    mp_api_key: str | None = None,
    nvidia_api_key: str | None = None,
    enable_ai: bool = False,
) -> GenericResult:
    source = Path(path)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    sample_id = sample_id or source.stem
    technique_name = detect_technique(source, technique)
    source_copy = out / source.name
    if source.resolve() != source_copy.resolve():
        shutil.copy2(source, source_copy)

    if technique_name == "eis":
        metrics, plots, tables = analyze_eis(source_copy, out, dpi, style)
    elif technique_name == "dpv":
        metrics, plots, tables = analyze_dpv(source_copy, out, dpi, style)
    elif technique_name == "gcd":
        metrics, plots, tables = analyze_gcd(source_copy, out, dpi, style)
    elif technique_name == "raman":
        metrics, plots, tables = analyze_raman(source_copy, out, dpi, style)
    else:
        raise ValueError(f"Generic analyzer does not handle technique {technique_name!r}.")

    metrics["sample_id"] = sample_id
    metrics["source_name"] = source.name
    metrics["dpi"] = dpi
    metrics["style"] = style

    material_result = materials_project_lookup(material_query or source.stem, api_key=mp_api_key)
    metrics["materials_project"] = {"ok": material_result.ok, "message": material_result.message, "data": material_result.data}
    if material_result.ok and material_result.data:
        pd.DataFrame(material_result.data).to_csv(out / "materials_project_candidates.csv", index=False)
        tables["Materials_Project"] = pd.DataFrame(material_result.data)

    if enable_ai:
        ai = nvidia_nim_commentary(metrics, api_key=nvidia_api_key)
        metrics["ai_commentary"] = {"ok": ai.ok, "message": ai.message, "data": ai.data}
        if ai.ok and ai.data:
            (out / "ai_commentary.md").write_text(ai.data["text"], encoding="utf-8")
    else:
        metrics["ai_commentary"] = {"ok": False, "message": "AI commentary disabled for this run."}

    _write_json(out / "analysis_summary.json", metrics)
    workbook = build_generic_workbook(out, source.name, metrics, plots, tables)
    _zip_dir(out, out / f"{sample_id}_all_outputs.zip")
    return GenericResult(metrics["technique"], sample_id, source.name, out, metrics, plots, workbook)
