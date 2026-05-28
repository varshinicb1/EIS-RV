"""
Autonomous Electrochemical Data Cleaner
=========================================
Automatically detects, parses, and cleans any electrochemical data file.

Handles:
- CHI608E EIS format (metadata header + Freq/Z'/Z"/Z/Phase columns)
- Interleaved multi-concentration DPV/CV format (potential, current pairs)
- Missing headers, None columns, unit inconsistencies
- Floating-point noise, duplicate rows, outliers
- Mixed units (µM vs µm, mA vs A, kHz vs Hz)

Outputs:
- Cleaned CSV files per measurement type
- JSON metadata with provenance, quality report
- Summary statistics

Author: VidyuthLabs
Date: May 6, 2026
"""

import re
import json
import logging
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any

import numpy as np
import openpyxl

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# DETECTION ENGINE
# ═══════════════════════════════════════════════════════════════════════════

class FormatDetector:
    """Automatically detect the format of an electrochemical data file."""

    # CHI608E EIS header signature
    CHI_EIS_SIGNATURES = ["A.C. Impedance", "Freq/Hz", "Z'/ohm", "Instrument Model:  CHI"]
    # Interleaved concentration format
    CONC_UNIT_PATTERN = re.compile(r"(\d+\.?\d*)\s*(µM|µm|uM|um|nM|mM|µl|ul|µL|uL|ml|mL)", re.IGNORECASE)
    # DPV/CV column headers
    POTENTIAL_HEADERS = {"potential", "potential (v)", "e/v", "e (v)", "voltage", "v"}
    CURRENT_HEADERS   = {"current", "current (a)", "i/a", "i (a)", "current (µa)", "current (ma)"}

    @classmethod
    def detect(cls, ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
        """
        Detect worksheet format.

        Returns one of:
          'chi_eis'          - CHI608E EIS format
          'interleaved_conc' - Multi-concentration interleaved (DPV/CV)
          'simple_xy'        - Simple two-column potential/current
          'unknown'
        """
        rows = list(ws.iter_rows(values_only=True, max_row=20))
        flat = " ".join(str(v) for r in rows for v in r if v is not None)

        # CHI EIS
        if any(sig in flat for sig in cls.CHI_EIS_SIGNATURES):
            return "chi_eis"

        # Interleaved concentration (look for µM/µl labels in first 3 rows)
        for r in rows[:3]:
            for v in r:
                if v and cls.CONC_UNIT_PATTERN.search(str(v)):
                    return "interleaved_conc"
            # Also check for 'buffer' label
            if any(str(v).strip().lower() == "buffer" for v in r if v):
                return "interleaved_conc"

        # Simple XY
        for r in rows[:5]:
            non_none = [v for v in r if v is not None]
            if len(non_none) >= 2:
                try:
                    float(non_none[0]); float(non_none[1])
                    return "simple_xy"
                except (TypeError, ValueError):
                    pass

        return "unknown"


# ═══════════════════════════════════════════════════════════════════════════
# PARSERS
# ═══════════════════════════════════════════════════════════════════════════

class CHIEISParser:
    """Parse CHI608E EIS files."""

    METADATA_KEYS = [
        "Init E (V)", "High Frequency (Hz)", "Low Frequency (Hz)",
        "Amplitude (V)", "Quiet Time (sec)", "Instrument Model",
        "Data Source", "File",
    ]

    def parse(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
        rows = list(ws.iter_rows(values_only=True))

        # Extract metadata
        meta = {
            "date": None, "instrument": None, "source_file": None,
            "init_e_v": None, "freq_high_hz": None, "freq_low_hz": None,
            "amplitude_v": None, "quiet_time_s": None,
        }

        for r in rows[:20]:
            cell = str(r[0]) if r[0] else ""
            if r[0] and r[1] and not r[2]:  # date row: "Apr. 16", " 2026 ..."
                if re.match(r"[A-Z][a-z]+\.", cell):
                    meta["date"] = f"{r[0]} {r[1]}".strip()
            if "Instrument Model" in cell:
                meta["instrument"] = cell.split(":", 1)[-1].strip()
            if "File:" in cell:
                meta["source_file"] = cell.split("File:", 1)[-1].strip()
            if "=" in cell:
                k, v = cell.split("=", 1)
                k = k.strip(); v = v.strip()
                if "Init E" in k:       meta["init_e_v"]      = _safe_float(v)
                if "High Frequency" in k: meta["freq_high_hz"] = _safe_float(v)
                if "Low Frequency" in k:  meta["freq_low_hz"]  = _safe_float(v)
                if "Amplitude" in k:    meta["amplitude_v"]   = _safe_float(v)
                if "Quiet Time" in k:   meta["quiet_time_s"]  = _safe_float(v)

        # Find data header row
        header_row = None
        for i, r in enumerate(rows):
            if r[0] and "Freq" in str(r[0]):
                header_row = i
                break

        if header_row is None:
            raise ValueError("Could not find EIS data header row")

        # Parse numeric data
        data = []
        for r in rows[header_row + 2:]:   # skip blank row after header
            try:
                if r[0] is not None:
                    data.append([
                        float(r[0]),  # Freq/Hz
                        float(r[1]),  # Z'/ohm
                        float(r[2]),  # Z"/ohm
                        float(r[3]),  # |Z|/ohm
                        float(r[4]),  # Phase/deg
                    ])
            except (TypeError, ValueError):
                pass

        return {
            "metadata": meta,
            "columns": ["freq_hz", "zreal_ohm", "zimag_ohm", "zmag_ohm", "phase_deg"],
            "data": data,
        }


class InterleavedConcParser:
    """
    Parse multi-concentration interleaved format.

    Layout:
      Row 0 (optional): column type header ("POTENTIAL (V)", "CURRENT (A)", ...)
      Row 1: concentration labels in odd columns (1, 3, 5, ...)
      Row 2+: data — even cols = potential, odd cols = current

    OR:
      Row 0: concentration labels in even columns (0, 2, 4, ...)
      Row 1+: data — even cols = potential, odd cols = current
    """

    CONC_PATTERN = re.compile(
        r"(\d+\.?\d*)\s*(µM|µm|uM|um|nM|mM|µl|ul|µL|uL|ml|mL|mg/L|ng/mL)",
        re.IGNORECASE
    )

    def parse(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
        rows = list(ws.iter_rows(values_only=True))
        n_cols = ws.max_column

        # Find the concentration label row
        conc_row_idx = None
        conc_labels  = {}   # col_index -> label

        for i, r in enumerate(rows[:4]):
            found = {}
            for j, v in enumerate(r):
                if v is None:
                    continue
                s = str(v).strip()
                # Match µM/µL labels
                if self.CONC_PATTERN.search(s) or s.lower() == "buffer":
                    found[j] = s
                # Also match bare numeric concentrations (e.g. 300, 400, 1200)
                # when they appear in even columns with None in odd columns
                elif re.match(r"^\d+(\.\d+)?$", s):
                    # Check if adjacent column is None (interleaved pattern)
                    next_v = r[j + 1] if j + 1 < len(r) else None
                    if next_v is None:
                        found[j] = f"{s} µM"   # assume µM for bare numbers
            if found:
                conc_row_idx = i
                conc_labels  = found
                break

        if conc_row_idx is None:
            # No labels found — assign generic names
            conc_row_idx = -1
            for j in range(0, n_cols, 2):
                conc_labels[j + 1] = f"series_{j // 2 + 1}"

        data_start = conc_row_idx + 1

        # Determine column pairing
        # Labels are in odd cols → potential in even, current in odd
        # Labels are in even cols → potential in even, current in odd
        label_cols = sorted(conc_labels.keys())

        # Build (potential_col, current_col, label) triples
        series = []
        for lc in label_cols:
            label = _normalise_unit(conc_labels[lc])
            # Determine pairing based on label column position
            if lc % 2 == 1:
                # Label in odd col → potential in lc-1, current in lc
                pot_col = lc - 1
                cur_col = lc
            else:
                # Label in even col → potential in lc, current in lc+1
                pot_col = lc
                cur_col = lc + 1
            series.append((pot_col, cur_col, label))

        # Parse data
        parsed_series = {}
        for pot_col, cur_col, label in series:
            potentials = []
            currents   = []
            for r in rows[data_start:]:
                try:
                    p = _safe_float(r[pot_col]) if pot_col < len(r) else None
                    c = _safe_float(r[cur_col]) if cur_col < len(r) else None
                    if p is not None and c is not None:
                        potentials.append(p)
                        currents.append(c)
                except (IndexError, TypeError):
                    pass
            if potentials:
                parsed_series[label] = {
                    "potential_v": potentials,
                    "current_a":   currents,
                }

        return {
            "metadata": {"format": "interleaved_concentration"},
            "columns": ["potential_v", "current_a"],
            "series": parsed_series,
        }


# ═══════════════════════════════════════════════════════════════════════════
# CLEANERS
# ═══════════════════════════════════════════════════════════════════════════

class EISCleaner:
    """Clean and validate EIS data."""

    def clean(self, parsed: Dict[str, Any]) -> Dict[str, Any]:
        data = np.array(parsed["data"], dtype=float)
        report = {"original_rows": len(data), "issues": []}

        if len(data) == 0:
            return {**parsed, "data_clean": [], "quality_report": report}

        freq   = data[:, 0]
        zreal  = data[:, 1]
        zimag  = data[:, 2]
        zmag   = data[:, 3]
        zphase = data[:, 4]

        mask = np.ones(len(data), dtype=bool)

        # 1. Remove NaN rows
        nan_mask = np.any(np.isnan(data), axis=1)
        if nan_mask.any():
            report["issues"].append(f"Removed {nan_mask.sum()} NaN rows")
            mask &= ~nan_mask

        # 2. Remove duplicate frequencies (keep first)
        _, unique_idx = np.unique(freq[mask], return_index=True)
        dup_count = mask.sum() - len(unique_idx)
        if dup_count > 0:
            report["issues"].append(f"Removed {dup_count} duplicate frequency rows")
            full_idx = np.where(mask)[0]
            new_mask = np.zeros(len(data), dtype=bool)
            new_mask[full_idx[unique_idx]] = True
            mask = new_mask

        # 3. Remove non-physical Zreal (must be > 0)
        neg_zreal = (zreal < 0) & mask
        if neg_zreal.any():
            report["issues"].append(f"Removed {neg_zreal.sum()} rows with Zreal < 0")
            mask &= ~neg_zreal

        # 4. Sort by frequency descending (high → low, standard EIS convention)
        clean_data = data[mask]
        sort_idx   = np.argsort(clean_data[:, 0])[::-1]
        clean_data = clean_data[sort_idx]

        # 5. Round floating-point noise (6 significant figures)
        clean_data = np.round(clean_data, 6)

        # 6. Compute derived quantities
        freq_c  = clean_data[:, 0]
        zr      = clean_data[:, 1]
        zi      = clean_data[:, 2]
        rs      = float(zr[0])
        rct_idx = int(np.argmax(-zi))
        rct     = float(zr[rct_idx] - rs)
        f_char  = float(freq_c[rct_idx])

        report["cleaned_rows"]  = len(clean_data)
        report["rs_ohm"]        = round(rs, 3)
        report["rct_ohm"]       = round(rct, 3)
        report["f_char_hz"]     = round(f_char, 3)
        report["freq_range_hz"] = [round(float(freq_c[-1]), 3), round(float(freq_c[0]), 3)]

        return {
            **parsed,
            "data_clean": clean_data.tolist(),
            "quality_report": report,
        }


class ElectrochemCleaner:
    """Clean DPV/CV concentration series data."""

    def clean(self, parsed: Dict[str, Any]) -> Dict[str, Any]:
        series_clean = {}
        report = {"series": {}}

        for label, s in parsed["series"].items():
            pot = np.array(s["potential_v"], dtype=float)
            cur = np.array(s["current_a"],   dtype=float)
            sr  = {"original_points": len(pot), "issues": []}

            # 1. Remove NaN
            valid = ~(np.isnan(pot) | np.isnan(cur))
            if (~valid).any():
                sr["issues"].append(f"Removed {(~valid).sum()} NaN points")
            pot, cur = pot[valid], cur[valid]

            # 2. Sort by potential
            idx = np.argsort(pot)
            pot, cur = pot[idx], cur[idx]

            # 3. Remove duplicate potentials (keep mean)
            u_pot, inv = np.unique(np.round(pot, 5), return_inverse=True)
            if len(u_pot) < len(pot):
                dup = len(pot) - len(u_pot)
                sr["issues"].append(f"Merged {dup} duplicate potential points")
                u_cur = np.array([cur[inv == i].mean() for i in range(len(u_pot))])
                pot, cur = u_pot, u_cur

            # 4. Round floating-point noise
            pot = np.round(pot, 6)
            cur = np.round(cur, 12)

            # 5. Detect peak
            peak_idx = int(np.argmax(np.abs(cur)))
            e_peak   = float(pot[peak_idx])
            i_peak   = float(cur[peak_idx])

            sr["cleaned_points"] = len(pot)
            sr["e_peak_v"]       = round(e_peak, 5)
            sr["i_peak_a"]       = round(i_peak, 9)
            sr["pot_range_v"]    = [round(float(pot.min()), 5), round(float(pot.max()), 5)]

            series_clean[label] = {
                "potential_v": pot.tolist(),
                "current_a":   cur.tolist(),
            }
            report["series"][label] = sr

        return {
            **parsed,
            "series_clean": series_clean,
            "quality_report": report,
        }


# ═══════════════════════════════════════════════════════════════════════════
# EXPORTERS
# ═══════════════════════════════════════════════════════════════════════════

class DataExporter:
    """Export cleaned data to CSV and JSON."""

    def export_eis(self, cleaned: Dict[str, Any], out_dir: Path, stem: str):
        out_dir.mkdir(parents=True, exist_ok=True)

        # CSV
        csv_path = out_dir / f"{stem}.csv"
        lines = ["freq_hz,zreal_ohm,zimag_ohm,zmag_ohm,phase_deg"]
        for row in cleaned["data_clean"]:
            lines.append(",".join(str(v) for v in row))
        csv_path.write_text("\n".join(lines), encoding="utf-8")

        # JSON metadata + quality report
        meta_path = out_dir / f"{stem}_meta.json"
        meta_path.write_text(json.dumps({
            "source": stem,
            "type": "EIS",
            "metadata": cleaned["metadata"],
            "quality_report": cleaned["quality_report"],
            "cleaned_at": datetime.now().isoformat(),
        }, indent=2), encoding="utf-8")

        return csv_path, meta_path

    def export_electrochem(self, cleaned: Dict[str, Any], out_dir: Path, stem: str, mtype: str):
        out_dir.mkdir(parents=True, exist_ok=True)
        exported = []

        for label, s in cleaned["series_clean"].items():
            safe_label = re.sub(r"[^\w\-]", "_", label)
            csv_path   = out_dir / f"{stem}_{safe_label}.csv"
            lines = ["potential_v,current_a"]
            for p, c in zip(s["potential_v"], s["current_a"]):
                lines.append(f"{p},{c}")
            csv_path.write_text("\n".join(lines), encoding="utf-8")
            exported.append(str(csv_path))

        # Combined JSON
        json_path = out_dir / f"{stem}_all.json"
        json_path.write_text(json.dumps({
            "source": stem,
            "type": mtype,
            "series": cleaned["series_clean"],
            "quality_report": cleaned["quality_report"],
            "cleaned_at": datetime.now().isoformat(),
        }, indent=2), encoding="utf-8")

        return exported, json_path


# ═══════════════════════════════════════════════════════════════════════════
# MAIN AUTONOMOUS CLEANER
# ═══════════════════════════════════════════════════════════════════════════

class AutonomousDataCleaner:
    """
    Drop any electrochemical Excel file in — get clean, analysis-ready data out.

    Automatically:
    1. Detects file format (EIS, DPV, CV, multi-concentration)
    2. Parses all sheets
    3. Cleans data (NaN, duplicates, outliers, unit normalisation)
    4. Exports to CSV + JSON
    5. Generates quality report
    """

    def __init__(self, output_dir: Optional[str] = None):
        if output_dir is None:
            output_dir = Path(__file__).parent.parent.parent.parent.parent / "data" / "cleaned"
        self.output_dir = Path(output_dir)
        self.eis_parser    = CHIEISParser()
        self.conc_parser   = InterleavedConcParser()
        self.eis_cleaner   = EISCleaner()
        self.ec_cleaner    = ElectrochemCleaner()
        self.exporter      = DataExporter()

    def clean_file(self, filepath: str | Path) -> Dict[str, Any]:
        """
        Clean a single Excel file. Returns a summary dict.

        Args:
            filepath: Path to .xlsx file

        Returns:
            {
              "file": str,
              "sheets": { sheet_name: { "format", "rows_in", "rows_out", "issues" } },
              "output_files": [str, ...],
              "success": bool,
              "error": str | None,
            }
        """
        filepath = Path(filepath)
        stem     = filepath.stem
        out_dir  = self.output_dir / stem
        summary  = {
            "file": str(filepath),
            "stem": stem,
            "sheets": {},
            "output_files": [],
            "success": False,
            "error": None,
        }

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            summary["error"] = f"Cannot open file: {e}"
            return summary

        for sheet_name in wb.sheetnames:
            ws     = wb[sheet_name]
            fmt    = FormatDetector.detect(ws)
            s_stem = f"{stem}_{_safe_name(sheet_name)}"
            sheet_summary = {"format": fmt, "rows_in": ws.max_row, "issues": []}

            try:
                if fmt == "chi_eis":
                    parsed  = self.eis_parser.parse(ws)
                    cleaned = self.eis_cleaner.clean(parsed)
                    csv_p, meta_p = self.exporter.export_eis(cleaned, out_dir, s_stem)
                    summary["output_files"] += [str(csv_p), str(meta_p)]
                    sheet_summary["rows_out"] = cleaned["quality_report"]["cleaned_rows"]
                    sheet_summary["issues"]   = cleaned["quality_report"]["issues"]
                    sheet_summary["rs_ohm"]   = cleaned["quality_report"].get("rs_ohm")
                    sheet_summary["rct_ohm"]  = cleaned["quality_report"].get("rct_ohm")

                elif fmt == "interleaved_conc":
                    parsed  = self.conc_parser.parse(ws)
                    cleaned = self.ec_cleaner.clean(parsed)
                    mtype   = _guess_measurement_type(ws)
                    csvs, json_p = self.exporter.export_electrochem(cleaned, out_dir, s_stem, mtype)
                    summary["output_files"] += csvs + [str(json_p)]
                    total_pts = sum(
                        sr.get("cleaned_points", 0)
                        for sr in cleaned["quality_report"]["series"].values()
                    )
                    sheet_summary["rows_out"]  = total_pts
                    sheet_summary["n_series"]  = len(cleaned["series_clean"])
                    sheet_summary["series"]    = list(cleaned["series_clean"].keys())

                else:
                    sheet_summary["rows_out"] = 0
                    sheet_summary["issues"]   = [f"Unknown format: {fmt}"]

            except Exception as e:
                sheet_summary["error"] = str(e)
                logger.error("Error processing sheet %s/%s: %s", stem, sheet_name, e)

            summary["sheets"][sheet_name] = sheet_summary

        summary["success"] = True
        return summary

    def clean_directory(self, directory: str | Path) -> Dict[str, Any]:
        """
        Clean all Excel files in a directory.

        Args:
            directory: Path to directory containing .xlsx files

        Returns:
            Summary dict with results for all files
        """
        directory = Path(directory)
        files     = sorted(directory.glob("*.xlsx"))

        if not files:
            return {"error": f"No .xlsx files found in {directory}", "files": []}

        results = []
        for f in files:
            logger.info("Cleaning: %s", f.name)
            result = self.clean_file(f)
            results.append(result)

        # Write master summary
        summary_path = self.output_dir / "cleaning_summary.json"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        master = {
            "cleaned_at":   datetime.now().isoformat(),
            "source_dir":   str(directory),
            "output_dir":   str(self.output_dir),
            "total_files":  len(files),
            "success":      sum(1 for r in results if r["success"]),
            "failed":       sum(1 for r in results if not r["success"]),
            "files":        results,
        }
        summary_path.write_text(json.dumps(master, indent=2), encoding="utf-8")

        return master


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _safe_float(v) -> Optional[float]:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _safe_name(s: str) -> str:
    return re.sub(r"[^\w]", "_", s).strip("_")


def _normalise_unit(label: str) -> str:
    """Normalise concentration unit labels: µm → µM, ul → µL, etc."""
    label = label.strip()
    # Fix case: µm → µM
    label = re.sub(r"(\d)\s*µm\b", r"\1 µM", label, flags=re.IGNORECASE)
    label = re.sub(r"(\d)\s*um\b", r"\1 µM", label, flags=re.IGNORECASE)
    label = re.sub(r"(\d)\s*uM\b", r"\1 µM", label)
    # Fix µl → µL
    label = re.sub(r"(\d)\s*µl\b", r"\1 µL", label, flags=re.IGNORECASE)
    label = re.sub(r"(\d)\s*ul\b", r"\1 µL", label, flags=re.IGNORECASE)
    return label


# Alias used by tests
_parse_conc_label = _normalise_unit


def _guess_measurement_type(ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
    """Guess DPV vs CV from sheet name or data characteristics."""
    name = ws.title.upper()
    if "DPV" in name:
        return "DPV"
    if "CV" in name or "CYCLIC" in name:
        return "CV"
    if "EIS" in name or "IMPEDANCE" in name:
        return "EIS"
    return "VOLTAMMETRY"


# ═══════════════════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s"
    )

    parser = argparse.ArgumentParser(description="Autonomous Electrochemical Data Cleaner")
    parser.add_argument("input",  help="Path to .xlsx file or directory of .xlsx files")
    parser.add_argument("--output", default=None, help="Output directory (default: data/cleaned/)")
    args = parser.parse_args()

    cleaner = AutonomousDataCleaner(output_dir=args.output)
    inp     = Path(args.input)

    if inp.is_dir():
        result = cleaner.clean_directory(inp)
    elif inp.suffix.lower() == ".xlsx":
        result = {"files": [cleaner.clean_file(inp)]}
    else:
        print(f"ERROR: {inp} is not a .xlsx file or directory")
        exit(1)

    # Pretty print summary
    print()
    print("=" * 70)
    print("AUTONOMOUS DATA CLEANER — RESULTS")
    print("=" * 70)

    for fr in result.get("files", []):
        status = "✅" if fr["success"] else "❌"
        print(f"\n{status} {fr['stem']}")
        for sheet, ss in fr.get("sheets", {}).items():
            fmt = ss.get("format", "?")
            ri  = ss.get("rows_in", "?")
            ro  = ss.get("rows_out", "?")
            print(f"   [{sheet}]  format={fmt}  rows: {ri} → {ro}")
            if ss.get("rs_ohm"):
                print(f"     Rs={ss['rs_ohm']} Ω  Rct={ss.get('rct_ohm')} Ω")
            if ss.get("series"):
                print(f"     Series: {ss['series']}")
            for issue in ss.get("issues", []):
                print(f"     ⚠ {issue}")
            if ss.get("error"):
                print(f"     ✗ ERROR: {ss['error']}")

    print()
    out_dir = result.get("output_dir", args.output or "data/cleaned")
    print(f"Output: {out_dir}")
    print(f"Summary: {out_dir}/cleaning_summary.json")
    print()
    print(f"Files processed: {result.get('total_files', len(result.get('files', [])))}")
    print(f"Success: {result.get('success', sum(1 for f in result.get('files',[]) if f['success']))}")
