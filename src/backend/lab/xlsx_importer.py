"""
XLSX importer for AnalyteX-style supercap workbooks.

The script ``scripts/import_agv_xlsx.py`` was the original entry point;
this module exposes the same logic as a callable so the FastAPI route
``POST /api/v2/lab/datasets/{id}/import/xlsx`` can use it.

Expected workbook layout
------------------------
* ``CV``  sheet: column F = Potential (V); subsequent numeric-header
                  columns = current (A) at the labelled scan rate (mV/s).
* ``GCD`` sheet: column C = Time (s); subsequent numeric-header columns
                  = voltage (V) for the labelled cycle.
* ``EIS`` sheet: columns E, F = Z' (Ω), Z'' (Ω) — frequency assumed
                  log-spaced ``f_max → f_min`` if not otherwise supplied.

What the importer extracts
--------------------------
For every CV scan rate, every GCD cycle, and the EIS spectrum, we
compute derived properties (ipa, ipc, dEp, capacitance, Rs, Rct, …)
plus a Randles-Sevcik linear fit ``ip ∝ √v``. The raw arrays go into
each row's ``conditions`` so the supercap analyzer can re-process them.
"""
from __future__ import annotations

import io
import math
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl


# ---- helpers --------------------------------------------------------


def _is_num(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ---- per-sheet extraction -----------------------------------------


def extract_cv(rows: list[tuple]) -> list[dict[str, Any]]:
    header = rows[0]
    pot_col = next(
        (i for i, v in enumerate(header)
         if isinstance(v, str) and "potential" in v.lower()),
        5,
    )
    sr_cols = [(i, v) for i, v in enumerate(header) if _is_num(v)]
    if not sr_cols:
        raise ValueError("CV sheet has no numeric scan-rate columns")

    data = [r for r in rows[1:]
            if pot_col < len(r) and r[pot_col] is not None and _is_num(r[pot_col])]
    if not data:
        raise ValueError("CV sheet has no numeric potential rows")

    pots = [r[pot_col] for r in data]
    out: list[dict[str, Any]] = []
    for col_idx, scan_mV_s in sr_cols:
        currents = []
        for r in data:
            v = r[col_idx] if col_idx < len(r) else None
            currents.append(v if _is_num(v) else float("nan"))

        if all(math.isnan(c) for c in currents):
            continue

        i_max_pot = max(range(len(pots)), key=lambda i: pots[i])
        forward_pots = pots[: i_max_pot + 1]
        forward_i    = currents[: i_max_pot + 1]
        reverse_pots = pots[i_max_pot:]
        reverse_i    = currents[i_max_pot:]

        ipa_idx = max(
            range(len(forward_i)),
            key=lambda i: forward_i[i] if not math.isnan(forward_i[i]) else -1e30,
        )
        i_pa, e_pa = forward_i[ipa_idx], forward_pots[ipa_idx]
        ipc_idx = min(
            range(len(reverse_i)),
            key=lambda i: reverse_i[i] if not math.isnan(reverse_i[i]) else 1e30,
        )
        i_pc, e_pc = reverse_i[ipc_idx], reverse_pots[ipc_idx]

        dEp = abs(e_pa - e_pc)
        ratio = abs(i_pa / i_pc) if i_pc not in (0, None) else None

        out.append({
            "scan_rate_mV_s": float(scan_mV_s),
            "scan_rate_V_s":  float(scan_mV_s) / 1000.0,
            "ipa_A":          float(i_pa),
            "ipc_A":          float(i_pc),
            "ipa_uA":         float(i_pa) * 1e6,
            "ipc_uA":         float(i_pc) * 1e6,
            "E_pa_V":         float(e_pa),
            "E_pc_V":         float(e_pc),
            "dEp_mV":         float(dEp) * 1000.0,
            "peak_current_ratio": ratio,
            "n_points":       len(pots),
            "potential_window_V": [min(pots), max(pots)],
            "raw_potential_V": list(pots),
            "raw_current_A":   list(currents),
        })
    return out


def randles_sevcik_fit(cv_rows: list[dict[str, Any]]) -> dict[str, Any]:
    if len(cv_rows) < 3:
        return {"available": False, "reason": "need ≥3 scan rates"}
    xs = [math.sqrt(r["scan_rate_V_s"]) for r in cv_rows]
    ys = [abs(r["ipa_A"]) for r in cv_rows]
    n = len(xs)
    sx = sum(xs); sy = sum(ys); sxx = sum(x * x for x in xs)
    sxy = sum(x * y for x, y in zip(xs, ys))
    denom = n * sxx - sx * sx
    if abs(denom) < 1e-30:
        return {"available": False, "reason": "degenerate"}
    slope = (n * sxy - sx * sy) / denom
    intercept = (sy - slope * sx) / n
    y_mean = sy / n
    ss_res = sum((y - (slope * x + intercept)) ** 2 for x, y in zip(xs, ys))
    ss_tot = sum((y - y_mean) ** 2 for y in ys)
    r2 = 1 - (ss_res / ss_tot) if ss_tot > 0 else None
    return {
        "available":            True,
        "slope_A_per_sqrt_Vs":  slope,
        "intercept_A":          intercept,
        "r_squared":            r2,
        "n_points":             n,
    }


def extract_gcd(rows: list[tuple], applied_current_A: float) -> list[dict[str, Any]]:
    header = rows[0]
    time_col = next(
        (i for i, v in enumerate(header)
         if isinstance(v, str) and "time" in v.lower()),
        2,
    )
    cycle_cols = [(i, v) for i, v in enumerate(header) if _is_num(v)]

    data = [r for r in rows[1:]
            if time_col < len(r) and r[time_col] is not None and _is_num(r[time_col])]
    times = [r[time_col] for r in data]

    out = []
    for col_idx, cycle_label in cycle_cols:
        voltages = [r[col_idx] if col_idx < len(r) else None for r in data]
        voltages = [v if _is_num(v) else float("nan") for v in voltages]
        clean = [(t, v) for t, v in zip(times, voltages) if not math.isnan(v)]
        if not clean:
            continue
        v_max = max(v for _, v in clean)
        v_min = min(v for _, v in clean)

        # Steepest discharge over a 20-point window (legacy heuristic; the
        # supercap analyzer does a proper linear-segment fit later).
        window = 20
        steepest = 0.0
        for i in range(len(clean) - window):
            t0, v0 = clean[i]; t1, v1 = clean[i + window]
            dt = t1 - t0
            if dt <= 0:
                continue
            slope = (v1 - v0) / dt
            if slope < steepest:
                steepest = slope
        cap_F = abs(applied_current_A / steepest) if steepest else None

        out.append({
            "cycle":               int(cycle_label),
            "applied_current_mA":  float(applied_current_A) * 1000.0,
            "v_max_V":             float(v_max),
            "v_min_V":             float(v_min),
            "v_swing_V":           float(v_max - v_min),
            "steepest_discharge_V_per_s": float(steepest),
            "capacitance_F":       cap_F,
            "duration_s":          float(clean[-1][0] - clean[0][0]),
            "n_points":            len(clean),
            "raw_time_s":          [t for t, _ in clean],
            "raw_voltage_V":       [v for _, v in clean],
        })
    return out


def extract_eis(
    rows: list[tuple],
    f_max_Hz: float,
    f_min_Hz: float,
) -> dict[str, Any]:
    data = [r for r in rows[1:] if 4 < len(r) and r[4] is not None and _is_num(r[4])]
    zr = [float(r[4]) for r in data]
    zi = [float(r[5]) for r in data if 5 < len(r)]
    zi = zi[: len(zr)]
    if len(zr) < 4 or len(zi) != len(zr):
        return {"available": False, "reason": "too few points"}

    n = len(zr)
    log_fmax = math.log10(f_max_Hz)
    log_fmin = math.log10(f_min_Hz)
    freqs = [10.0 ** (log_fmax - i * (log_fmax - log_fmin) / (n - 1))
             for i in range(n)]

    rs_zr = None
    for i in range(len(zi) - 1):
        if zi[i] * zi[i + 1] <= 0 and zi[i] != zi[i + 1]:
            t = zi[i] / (zi[i] - zi[i + 1])
            rs_zr = zr[i] + t * (zr[i + 1] - zr[i])
            break
    if rs_zr is None:
        rs_zr = zr[0]

    apex_idx = max(range(len(zi)), key=lambda i: -zi[i])
    apex_zr = zr[apex_idx]
    rct = max(0.0, 2.0 * (apex_zr - rs_zr))

    return {
        "available":         True,
        "n_points":          n,
        "f_max_Hz":          f_max_Hz,
        "f_min_Hz":          f_min_Hz,
        "Rs_ohm":            float(rs_zr),
        "Rct_ohm":           float(rct),
        "apex_Zreal_ohm":    float(apex_zr),
        "apex_Zimag_ohm":    float(zi[apex_idx]),
        "raw_frequency_Hz":  freqs,
        "raw_Z_real_ohm":    zr,
        "raw_Z_imag_ohm":    zi,
    }


# ---- top-level: bytes-in → rows-out ------------------------------


@dataclass
class XlsxImportOptions:
    material:           str = "AGV"
    electrolyte:        str = "unknown"
    gcd_current_mA:     float = 1.0
    eis_fmax_Hz:        float = 1.0e5
    eis_fmin_Hz:        float = 1.0e-2
    electrode_area_cm2: Optional[float] = None
    source_filename:    Optional[str] = None


@dataclass
class XlsxImportResult:
    rows:        list[dict[str, Any]]
    n_cv:        int = 0
    n_gcd:       int = 0
    has_eis:     bool = False
    rs_fit:      dict[str, Any] = field(default_factory=dict)


def import_xlsx_bytes(
    xlsx_bytes: bytes,
    opts: XlsxImportOptions,
) -> XlsxImportResult:
    """
    Parse the workbook in memory, run the per-sheet extractors, and
    return the list of rows ready for ``LabDatasetManager.add_rows``.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes),
                                 data_only=True, read_only=True)
    needed = {"CV", "GCD", "EIS"}
    missing = needed - set(wb.sheetnames)
    if missing:
        raise ValueError(
            f"workbook is missing sheet(s) {sorted(missing)}; "
            f"got {wb.sheetnames}"
        )

    cv_rows  = list(wb["CV"].iter_rows(values_only=True))
    gcd_rows = list(wb["GCD"].iter_rows(values_only=True))
    eis_rows = list(wb["EIS"].iter_rows(values_only=True))

    cv_results  = extract_cv(cv_rows)
    rs_fit      = randles_sevcik_fit(cv_results)
    gcd_results = extract_gcd(gcd_rows, opts.gcd_current_mA / 1000.0)
    eis_result  = extract_eis(eis_rows, opts.eis_fmax_Hz, opts.eis_fmin_Hz)

    common_conditions: dict[str, Any] = {
        "electrolyte":   opts.electrolyte,
        "source_file":   opts.source_filename,
        "instrument":    "AnalyteX (assumed)",
    }
    if opts.electrode_area_cm2 is not None:
        common_conditions["electrode_area_cm2"] = opts.electrode_area_cm2

    rows: list[dict[str, Any]] = []
    for r in cv_results:
        rows.append({
            "formula": opts.material,
            "name":    f"CV {int(r['scan_rate_mV_s'])} mV/s",
            "conditions": {
                **common_conditions,
                "experiment":         "CV",
                "scan_rate_mV_s":     r["scan_rate_mV_s"],
                "scan_rate_V_s":      r["scan_rate_V_s"],
                "potential_window_V": r["potential_window_V"],
                "raw_potential_V":    r["raw_potential_V"],
                "raw_current_A":      r["raw_current_A"],
            },
            "properties": {
                "ipa_uA":             r["ipa_uA"],
                "ipc_uA":             r["ipc_uA"],
                "dEp_mV":             r["dEp_mV"],
                "peak_current_ratio": r["peak_current_ratio"],
                "E_pa_V":             r["E_pa_V"],
                "E_pc_V":             r["E_pc_V"],
            },
            "notes":  (f"derived from {opts.source_filename or 'xlsx'} → CV sheet, "
                       f"scan rate {int(r['scan_rate_mV_s'])} mV/s"),
            "source": "lab",
        })

    if rs_fit.get("available"):
        rows.append({
            "formula": opts.material,
            "name":    "Randles-Sevcik fit (ip vs √v)",
            "conditions": {**common_conditions, "experiment": "CV-summary"},
            "properties": {
                "rs_slope_A_per_sqrt_Vs": rs_fit["slope_A_per_sqrt_Vs"],
                "rs_intercept_A":         rs_fit["intercept_A"],
                "rs_r_squared":           rs_fit["r_squared"],
            },
            "notes":  "ip ∝ √v ⇒ diffusion-controlled; slope encodes A·C·√D.",
            "source": "lab",
        })

    for r in gcd_results:
        rows.append({
            "formula": opts.material,
            "name":    f"GCD cycle {r['cycle']}",
            "conditions": {
                **common_conditions,
                "experiment":           "GCD",
                "cycle":                r["cycle"],
                "applied_current_mA":   r["applied_current_mA"],
                "raw_time_s":           r["raw_time_s"],
                "raw_voltage_V":        r["raw_voltage_V"],
            },
            "properties": {
                "capacitance_F":              r["capacitance_F"],
                "v_max_V":                    r["v_max_V"],
                "v_min_V":                    r["v_min_V"],
                "v_swing_V":                  r["v_swing_V"],
                "steepest_discharge_V_per_s": r["steepest_discharge_V_per_s"],
            },
            "notes":  (f"derived from {opts.source_filename or 'xlsx'} → GCD cycle "
                       f"{r['cycle']} at {r['applied_current_mA']:.2f} mA"),
            "source": "lab",
        })

    if eis_result.get("available"):
        e = eis_result
        rows.append({
            "formula": opts.material,
            "name":    "EIS Nyquist",
            "conditions": {
                **common_conditions,
                "experiment":         "EIS",
                "f_max_Hz":           e["f_max_Hz"],
                "f_min_Hz":           e["f_min_Hz"],
                "raw_frequency_Hz":   e["raw_frequency_Hz"],
                "raw_Z_real_ohm":     e["raw_Z_real_ohm"],
                "raw_Z_imag_ohm":     e["raw_Z_imag_ohm"],
            },
            "properties": {
                "rs_ohm":          e["Rs_ohm"],
                "rct_ohm":         e["Rct_ohm"],
                "apex_zreal_ohm":  e["apex_Zreal_ohm"],
                "apex_zimag_ohm":  e["apex_Zimag_ohm"],
            },
            "notes":  ("derived from EIS sheet; frequency assumed log-spaced "
                       f"from {e['f_max_Hz']:.0e} Hz to {e['f_min_Hz']:.0e} Hz."),
            "source": "lab",
        })

    return XlsxImportResult(
        rows=rows,
        n_cv=len(cv_results),
        n_gcd=len(gcd_results),
        has_eis=eis_result.get("available", False),
        rs_fit=rs_fit,
    )
