"""
Deep Analysis of FOG Lab Dataset
==================================
Analyzes all 6 Excel files in the fog dataset:
- 4 EIS files (Bare GCE, Ferric Oxide, FOG, rGO)
- 1 DPV file (FOG concentration study)
- 1 CV file (Gomutra concentration study)

Author: VidyuthLabs
Date: May 6, 2026
"""

import openpyxl
import numpy as np
from pathlib import Path

BASE = Path("Lab data/fog differet data/fog differet data")

# ═══════════════════════════════════════════════════════════════════════════
# EIS ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

EIS_FILES = {
    "Bare GCE":     "EIS BARE GCE.xlsx",
    "Ferric Oxide": "EIS FERRIC OXIDE.xlsx",
    "FOG":          "EIS FOG.xlsx",
    "rGO":          "EIS rGO.xlsx",
}

print("=" * 70)
print("EIS ANALYSIS  —  CHI608E  |  Ferricyanide redox probe")
print("=" * 70)

eis_data = {}
for label, fname in EIS_FILES.items():
    wb = openpyxl.load_workbook(BASE / fname, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Metadata (rows 0-15)
    meta = {}
    for r in rows[:16]:
        if r[0] and "=" in str(r[0]):
            k, v = str(r[0]).split("=", 1)
            meta[k.strip()] = v.strip()

    # Data starts at row index 17
    data = []
    for r in rows[17:]:
        try:
            if r[0] is not None:
                data.append([float(r[0]), float(r[1]), float(r[2]),
                              float(r[3]), float(r[4])])
        except (TypeError, ValueError):
            pass

    if not data:
        print(f"\n{label}: NO DATA FOUND")
        continue

    arr    = np.array(data)
    freq   = arr[:, 0]
    zreal  = arr[:, 1]
    zimag  = arr[:, 2]   # negative
    zmag   = arr[:, 3]
    zphase = arr[:, 4]

    rs          = zreal[0]                          # Rs at highest freq
    neg_zimag   = -zimag
    rct_idx     = int(np.argmax(neg_zimag))
    rct_approx  = zreal[rct_idx] - rs
    f_char      = float(freq[rct_idx])

    eis_data[label] = dict(
        freq=freq, zreal=zreal, zimag=zimag, zmag=zmag, zphase=zphase,
        rs=rs, rct=rct_approx, f_char=f_char,
        n_points=len(data), meta=meta,
        date=f"{rows[0][0]} {rows[0][1]}"
    )

    print(f"\n{label}")
    print(f"  Date:              {rows[0][0]} {rows[0][1]}")
    print(f"  Init E:            {meta.get('Init E (V)', '?')} V")
    print(f"  Amplitude:         {meta.get('Amplitude (V)', '?')} V")
    print(f"  Freq range:        {freq[-1]:.1f} – {freq[0]:.0f} Hz")
    print(f"  Points:            {len(data)}")
    print(f"  Rs (solution):     {rs:.1f} Ω")
    print(f"  Rct (charge xfer): ~{rct_approx:.0f} Ω")
    print(f"  f_char:            {f_char:.1f} Hz")
    print(f"  |Z| at 1 Hz:       {zmag[-1]:.0f} Ω")
    print(f"  Phase at 1 Hz:     {zphase[-1]:.1f}°")

print()
print("=" * 70)
print("EIS COMPARISON TABLE")
print("=" * 70)
print(f"{'Electrode':<20} {'Rs (Ω)':>10} {'Rct (Ω)':>12} {'f_char (Hz)':>14} {'Rct/Rs':>8}")
print("-" * 70)
for label, d in eis_data.items():
    ratio = d["rct"] / d["rs"] if d["rs"] > 0 else 0
    print(f"{label:<20} {d['rs']:>10.1f} {d['rct']:>12.0f} {d['f_char']:>14.1f} {ratio:>8.1f}")

# ═══════════════════════════════════════════════════════════════════════════
# DPV ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

print()
print("=" * 70)
print("DPV ANALYSIS  —  FOG concentration study")
print("=" * 70)

wb = openpyxl.load_workbook(BASE / "DPV FOG.xlsx", data_only=True)

# Sheet3 has the labelled concentration data
ws3 = wb["Sheet3"]
rows3 = list(ws3.iter_rows(values_only=True))

# Row 1 = concentration labels (odd cols = potential, even = current)
conc_labels = []
for v in rows3[1]:
    if v is not None:
        conc_labels.append(str(v))

print(f"\nSheet3 — Concentration series")
print(f"  Concentrations: {conc_labels}")
print(f"  Data rows: {ws3.max_row - 2}")

# Parse each concentration column pair
conc_data = {}
for col_pair_idx, label in enumerate(conc_labels):
    pot_col = col_pair_idx * 2
    cur_col = col_pair_idx * 2 + 1
    potentials = []
    currents   = []
    for r in rows3[2:]:
        try:
            p = float(r[pot_col]) if r[pot_col] is not None else None
            c = float(r[cur_col]) if r[cur_col] is not None else None
            if p is not None and c is not None:
                potentials.append(p)
                currents.append(c)
        except (TypeError, ValueError):
            pass
    if potentials:
        conc_data[label] = dict(potential=np.array(potentials),
                                current=np.array(currents))

print(f"\n  {'Concentration':<15} {'E_peak (V)':>12} {'I_peak (A)':>14} {'I_net (A)':>14}")
print("  " + "-" * 60)
for label, d in conc_data.items():
    # Find peak (max current)
    peak_idx = int(np.argmax(d["current"]))
    e_peak   = d["potential"][peak_idx]
    i_peak   = d["current"][peak_idx]
    i_net    = i_peak - d["current"][0]   # net above baseline
    print(f"  {label:<15} {e_peak:>12.4f} {i_peak:>14.5f} {i_net:>14.5f}")

# Sheet1 — multi-concentration DPV (columns: potential, current pairs)
ws1 = wb["Sheet1"]
rows1 = list(ws1.iter_rows(values_only=True))
conc_header = rows1[0]
print(f"\nSheet1 — Concentrations in header: {[v for v in conc_header if v is not None]}")
print(f"  Data rows: {ws1.max_row - 1}")

# ═══════════════════════════════════════════════════════════════════════════
# GOMUTRA CONCENTRATION STUDY (CV)
# ═══════════════════════════════════════════════════════════════════════════

print()
print("=" * 70)
print("GOMUTRA CONCENTRATION STUDY  —  CV")
print("=" * 70)

wb_g = openpyxl.load_workbook(BASE / "GOMUTRA CONCENTRATION STUDIES.xlsx", data_only=True)
ws_g = wb_g.active
rows_g = list(ws_g.iter_rows(values_only=True))

# Row 1 = concentration labels
conc_labels_g = [str(v) for v in rows_g[1] if v is not None]
print(f"\n  Concentrations: {conc_labels_g}")
print(f"  Data rows: {ws_g.max_row - 2}")
print(f"  Potential range: {rows_g[2][0]:.3f} to {rows_g[-1][0]:.3f} V")

# Parse each concentration
gomutra_data = {}
for col_pair_idx, label in enumerate(conc_labels_g):
    pot_col = col_pair_idx * 2
    cur_col = col_pair_idx * 2 + 1
    potentials = []
    currents   = []
    for r in rows_g[2:]:
        try:
            p = float(r[pot_col]) if r[pot_col] is not None else None
            c = float(r[cur_col]) if r[cur_col] is not None else None
            if p is not None and c is not None:
                potentials.append(p)
                currents.append(c)
        except (TypeError, ValueError):
            pass
    if potentials:
        gomutra_data[label] = dict(potential=np.array(potentials),
                                   current=np.array(currents))

print(f"\n  {'Concentration':<15} {'E_range (V)':>20} {'I_max (A)':>14} {'I_min (A)':>14}")
print("  " + "-" * 68)
for label, d in gomutra_data.items():
    e_range = f"{d['potential'].min():.3f} to {d['potential'].max():.3f}"
    print(f"  {label:<15} {e_range:>20} {d['current'].max():>14.3e} {d['current'].min():>14.3e}")

# ═══════════════════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════════════════

print()
print("=" * 70)
print("DATASET SUMMARY")
print("=" * 70)
print(f"""
Files:          6 Excel files
Instrument:     CHI608E (CH Instruments)
Experiment:     FOG (Ferric Oxide on Graphene) biosensor

EIS Files (4):
  - Bare GCE       : baseline electrode
  - Ferric Oxide   : Fe2O3 modified electrode
  - FOG            : Ferric Oxide + Graphene composite
  - rGO            : reduced Graphene Oxide

DPV File (1):
  - FOG electrode, concentration series
  - Concentrations: {list(conc_data.keys())}
  - Points per curve: {len(list(conc_data.values())[0]['potential']) if conc_data else 0}

Gomutra CV File (1):
  - Gomutra (cow urine) concentration study
  - Concentrations: {conc_labels_g}
  - Points per curve: {len(list(gomutra_data.values())[0]['potential']) if gomutra_data else 0}

Key Findings:
  EIS: Rct order = {' < '.join(sorted(eis_data.keys(), key=lambda k: eis_data[k]['rct']))}
  (Lower Rct = better electron transfer = better electrode performance)
""")
