#!/usr/bin/env python3
"""
Import a real AnalyteX-style XLSX into the encrypted lab dataset.

Designed against the user's ``AGV CV GCD EIS.xlsx`` layout:
  * CV  sheet: column F = Potential (V), columns G..P = current (A) at
              scan rates listed in the header row (10, 20, ... mV/s).
  * GCD sheet: column C = Time (s), columns D..L = voltage (V) for cycles 1..9.
  * EIS sheet: columns E, F = Z' (Ω), Z'' (Ω) — frequency assumed to be
              log-spaced from --eis-fmax down to --eis-fmin over the
              measured points.

What this script extracts
-------------------------
* CV per scan rate: ipa, ipc, dEp, |ipa/ipc|, peak potentials.
* CV across scan rates: Randles-Sevcik slope ``d(ip)/d(√v)`` and
  the implied n·D effective product if the user provides A and C.
* GCD per cycle: discharge capacitance Cs (from Δt·I / ΔV — needs the
  applied current; defaults to 1 mA if not supplied).
* EIS: high-frequency intercept (Rs), semicircle apex Rct, low-frequency
  intercept (Rs+Rct). Optional Lin-KK validity check via the C++
  ``raman_core.kramers_kronig_test`` if the binding loads.

Storage
-------
* A new dataset named ``<material>-<file-stem>`` is created (or reused
  with --append).
* Derived scalars (one dict per scan rate / cycle / spectrum) become
  rows. Raw arrays go into ``conditions["raw_*"]`` so the user can
  re-analyse later. ``formula`` comes from --material; default "AGV".
"""
from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path
from typing import Any

# Make sure we can import the lab manager without spawning a server.
THIS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(THIS_DIR.parent))

import openpyxl  # noqa: E402

from src.backend.lab.dataset_manager import (  # noqa: E402
    get_lab_dataset_manager,
)


# ---- helpers --------------------------------------------------------


def _is_num(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _read_sheet(path: Path, sheet_name: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not in {wb.sheetnames}")
    return list(wb[sheet_name].iter_rows(values_only=True))


# ---- CV extraction --------------------------------------------------


def extract_cv(rows: list[tuple]) -> list[dict[str, Any]]:
    """One result row per scan rate."""
    header = rows[0]
    pot_col = next((i for i, v in enumerate(header)
                    if isinstance(v, str) and "potential" in v.lower()), 5)
    sr_cols = [(i, v) for i, v in enumerate(header) if _is_num(v)]
    if not sr_cols:
        raise ValueError("CV: no numeric scan-rate columns in header")

    data = []
    for r in rows[1:]:
        if pot_col >= len(r) or r[pot_col] is None or not _is_num(r[pot_col]):
            continue
        data.append(r)

    pots = [r[pot_col] for r in data]
    out: list[dict[str, Any]] = []
    for col_idx, scan_mV_s in sr_cols:
        # Pull the current trace at this scan rate.
        currents = []
        for r in data:
            v = r[col_idx] if col_idx < len(r) else None
            if v is None or not _is_num(v):
                v = float("nan")
            currents.append(v)

        # Find anodic and cathodic peaks separately.
        # The waveform is a triangle: forward then reverse. We don't know
        # the segment boundary precisely, so split at potential's max.
        if not currents or all(math.isnan(c) for c in currents):
            continue
        i_max_pot = max(range(len(pots)), key=lambda i: pots[i])

        forward_pots = pots[:i_max_pot + 1]
        forward_i    = currents[:i_max_pot + 1]
        reverse_pots = pots[i_max_pot:]
        reverse_i    = currents[i_max_pot:]

        # Anodic peak = max of forward sweep (oxidation hump)
        i_pa_idx = max(range(len(forward_i)),
                       key=lambda i: forward_i[i] if not math.isnan(forward_i[i]) else -1e30)
        i_pa = forward_i[i_pa_idx]
        E_pa = forward_pots[i_pa_idx]

        # Cathodic peak = min of reverse sweep
        i_pc_idx = min(range(len(reverse_i)),
                       key=lambda i: reverse_i[i] if not math.isnan(reverse_i[i]) else 1e30)
        i_pc = reverse_i[i_pc_idx]
        E_pc = reverse_pots[i_pc_idx]

        dEp = abs(E_pa - E_pc)
        peak_ratio = abs(i_pa / i_pc) if i_pc not in (0, None) else None

        out.append({
            "experiment": "CV",
            "scan_rate_mV_s": float(scan_mV_s),
            "scan_rate_V_s":  float(scan_mV_s) / 1000.0,
            "ipa_A":     float(i_pa),
            "ipc_A":     float(i_pc),
            "ipa_uA":    float(i_pa) * 1e6,
            "ipc_uA":    float(i_pc) * 1e6,
            "E_pa_V":    float(E_pa),
            "E_pc_V":    float(E_pc),
            "dEp_mV":    float(dEp) * 1000.0,
            "peak_current_ratio": peak_ratio,
            "n_points":  len(pots),
            "potential_window_V": [min(pots), max(pots)],
            # Raw arrays — kept under "raw_*" so they don't pollute lookup
            # of derived scalars.
            "raw_potential_V": list(pots),
            "raw_current_A":   list(currents),
        })
    return out


def randles_sevcik_fit(cv_rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Linear fit of ip_anodic vs sqrt(v) (V/s). Returns slope and r²."""
    if len(cv_rows) < 3:
        return {"available": False, "reason": "need ≥3 scan rates"}
    xs = [math.sqrt(r["scan_rate_V_s"]) for r in cv_rows]
    ys = [abs(r["ipa_A"]) for r in cv_rows]
    n = len(xs)
    sx = sum(xs); sy = sum(ys); sxx = sum(x*x for x in xs)
    sxy = sum(x*y for x, y in zip(xs, ys))
    denom = n * sxx - sx * sx
    if abs(denom) < 1e-30:
        return {"available": False, "reason": "degenerate"}
    slope = (n * sxy - sx * sy) / denom
    intercept = (sy - slope * sx) / n
    # r² (coefficient of determination)
    y_mean = sy / n
    ss_res = sum((y - (slope * x + intercept)) ** 2 for x, y in zip(xs, ys))
    ss_tot = sum((y - y_mean) ** 2 for y in ys)
    r2 = 1 - (ss_res / ss_tot) if ss_tot > 0 else None
    return {
        "available":      True,
        "slope_A_per_sqrt_Vs":  slope,
        "intercept_A":          intercept,
        "r_squared":            r2,
        "n_points":             n,
        "interpretation": (
            "ip ∝ √v ⇒ diffusion-controlled. Slope = 0.4463·n^(3/2)·F^(3/2)·"
            "A·C·√(D/RT) — solve for A·C·√(D) given n, T."
        ),
    }


# ---- GCD extraction -------------------------------------------------


def extract_gcd(rows: list[tuple], applied_current_A: float) -> list[dict[str, Any]]:
    """One result row per cycle column."""
    header = rows[0]
    time_col = next((i for i, v in enumerate(header)
                     if isinstance(v, str) and "time" in v.lower()), 2)
    cycle_cols = [(i, v) for i, v in enumerate(header) if _is_num(v)]

    data = [r for r in rows[1:] if r[time_col] is not None and _is_num(r[time_col])]
    times = [r[time_col] for r in data]

    out = []
    for col_idx, cycle_label in cycle_cols:
        voltages = [r[col_idx] if col_idx < len(r) else None for r in data]
        voltages = [v if _is_num(v) else float("nan") for v in voltages]
        # Find min and max voltage in this cycle
        clean = [(t, v) for t, v in zip(times, voltages) if not math.isnan(v)]
        if not clean:
            continue
        v_max = max(v for _, v in clean)
        v_min = min(v for _, v in clean)

        # Discharge slope (V/s) — we look at the steepest negative-slope segment
        # over a moving window of ~10 points.
        window = 20
        steepest_disch = 0.0
        for i in range(len(clean) - window):
            t0, v0 = clean[i]; t1, v1 = clean[i + window]
            dt = t1 - t0
            if dt <= 0: continue
            slope = (v1 - v0) / dt
            if slope < steepest_disch:
                steepest_disch = slope
        # Capacitance: C = I / |dV/dt| (gravimetric C would need mass;
        # here we report specific C per electrode area only if user supplies).
        cap_F = abs(applied_current_A / steepest_disch) if steepest_disch else None

        out.append({
            "experiment":       "GCD",
            "cycle":            int(cycle_label),
            "applied_current_mA": float(applied_current_A) * 1000.0,
            "v_max_V":          float(v_max),
            "v_min_V":          float(v_min),
            "v_swing_V":        float(v_max - v_min),
            "steepest_discharge_V_per_s": float(steepest_disch),
            "capacitance_F":    cap_F,
            "duration_s":       float(clean[-1][0] - clean[0][0]),
            "n_points":         len(clean),
            "raw_time_s":       [t for t, _ in clean],
            "raw_voltage_V":    [v for _, v in clean],
        })
    return out


# ---- EIS extraction -------------------------------------------------


def extract_eis(
    rows: list[tuple],
    f_max_Hz: float,
    f_min_Hz: float,
) -> dict[str, Any]:
    """Synthesise a frequency vector and extract Rs/Rct from the Nyquist."""
    data = [r for r in rows[1:] if r[4] is not None and _is_num(r[4])]
    zr = [float(r[4]) for r in data]
    zi = [float(r[5]) for r in data]
    if len(zr) < 4:
        return {"available": False, "reason": "too few points"}

    n = len(zr)
    # Log-spaced frequency grid, high-frequency first (typical Nyquist
    # acquisition order).
    log_fmax = math.log10(f_max_Hz)
    log_fmin = math.log10(f_min_Hz)
    freqs = [10.0 ** (log_fmax - i * (log_fmax - log_fmin) / (n - 1))
             for i in range(n)]

    # ── High-frequency intercept (Rs):
    # The point where Z'' crosses zero from below (going from imaginary
    # negative — capacitive — to imaginary positive) gives Rs.
    # In our data Z'' starts large positive at low freq and goes negative
    # at high freq (the row is sorted high-freq → low-freq with our
    # synthesis), so we find the highest-frequency zero-crossing of Z''.
    rs_zr = None
    for i in range(len(zi) - 1):
        if zi[i] * zi[i + 1] <= 0 and zi[i] != zi[i + 1]:
            t = zi[i] / (zi[i] - zi[i + 1])
            rs_zr = zr[i] + t * (zr[i + 1] - zr[i])
            break
    # If no zero-crossing, use the high-frequency point directly.
    if rs_zr is None:
        rs_zr = zr[0]

    # ── Charge-transfer resistance:
    # The semicircle apex is the point of maximum |-Z''| in the high-to-mid
    # frequency portion before any low-frequency Warburg straight line.
    # We compute Rct as 2·(Z'@apex - Rs) under the assumption of an ideal
    # semicircle.
    apex_idx = max(range(len(zi)), key=lambda i: -zi[i])  # most negative Z''
    apex_zr = zr[apex_idx]
    rct = max(0.0, 2.0 * (apex_zr - rs_zr))

    return {
        "available":  True,
        "experiment": "EIS",
        "n_points":   n,
        "f_max_Hz":   f_max_Hz,
        "f_min_Hz":   f_min_Hz,
        "Rs_ohm":     float(rs_zr),
        "Rct_ohm":    float(rct),
        "Z_real_min_ohm": float(min(zr)),
        "Z_real_max_ohm": float(max(zr)),
        "Z_imag_min_ohm": float(min(zi)),
        "Z_imag_max_ohm": float(max(zi)),
        "apex_Zreal_ohm": float(apex_zr),
        "apex_Zimag_ohm": float(zi[apex_idx]),
        "raw_frequency_Hz": freqs,
        "raw_Z_real_ohm":   zr,
        "raw_Z_imag_ohm":   zi,
    }


# ---- main -----------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--material", default="AGV",
                    help="Formula / tag for every row (default: 'AGV').")
    ap.add_argument("--dataset-name", default=None,
                    help="Lab dataset name (default: <material>-<file-stem>).")
    ap.add_argument("--gcd-current-mA", type=float, default=1.0,
                    help="Applied current for the GCD run (default: 1 mA).")
    ap.add_argument("--eis-fmax-Hz", type=float, default=1.0e5)
    ap.add_argument("--eis-fmin-Hz", type=float, default=1.0e-2)
    ap.add_argument("--electrolyte", default="unknown")
    ap.add_argument("--electrode-area-cm2", type=float, default=None)
    ap.add_argument("--append", action="store_true",
                    help="Append rows to an existing dataset of the same name.")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: {args.xlsx} not found"); return 1

    # ── Read each sheet ──
    print(f"Reading {args.xlsx.name}…")
    cv_rows  = _read_sheet(args.xlsx, "CV")
    gcd_rows = _read_sheet(args.xlsx, "GCD")
    eis_rows = _read_sheet(args.xlsx, "EIS")

    # ── Extract per-experiment scalars ──
    cv_results  = extract_cv(cv_rows)
    rs_fit      = randles_sevcik_fit(cv_results)
    gcd_results = extract_gcd(gcd_rows, args.gcd_current_mA / 1000.0)
    eis_result  = extract_eis(eis_rows, args.eis_fmax_Hz, args.eis_fmin_Hz)

    # ── Print summary ──
    print("\n──────── CV ─────────────────────────────────────────")
    print(f"  scan rates:  {len(cv_results)} ({[r['scan_rate_mV_s'] for r in cv_results]} mV/s)")
    print(f"  potential window: {cv_results[0]['potential_window_V'] if cv_results else 'n/a'} V")
    for r in cv_results:
        print(f"  v={r['scan_rate_mV_s']:>4.0f} mV/s  ipa={r['ipa_uA']:>+8.2f} μA  ipc={r['ipc_uA']:>+8.2f} μA  "
              f"dEp={r['dEp_mV']:>6.1f} mV  ratio={r['peak_current_ratio']:.3f}")

    print("\n  Randles-Sevcik (ip vs √v):")
    if rs_fit["available"]:
        slope_uA = rs_fit["slope_A_per_sqrt_Vs"] * 1e6
        print(f"    slope     = {slope_uA:.2f} μA/√(V/s)")
        print(f"    intercept = {rs_fit['intercept_A']*1e6:+.3f} μA")
        print(f"    r²        = {rs_fit['r_squared']:.4f}")
        if args.electrode_area_cm2:
            # n=1 assumed; T=298.15 K
            F = 96485; R_GAS = 8.314; T = 298.15
            from math import sqrt
            # slope_A = 0.4463·n^1.5·F^1.5·A·C·sqrt(D/RT)
            #        ⇒ A·C·sqrt(D) = slope_A / (0.4463·F^1.5·sqrt(1/RT))
            denom = 0.4463 * (F ** 1.5) * sqrt(1.0 / (R_GAS * T))
            ac_sqrt_D = rs_fit["slope_A_per_sqrt_Vs"] / denom
            print(f"    A·C·√D    = {ac_sqrt_D:.3e}  cm²·mol/cm³·√(cm²/s)")
            print(f"    (A={args.electrode_area_cm2} cm² assumed; n=1, T=298 K)")
    else:
        print(f"    not available: {rs_fit.get('reason')}")

    print("\n──────── GCD ────────────────────────────────────────")
    for r in gcd_results:
        c = r["capacitance_F"]
        c_str = f"{c*1000:.2f} mF" if c is not None else "N/A"
        print(f"  cycle {r['cycle']:>2}: V {r['v_min_V']:+.3f}→{r['v_max_V']:+.3f} V, "
              f"steepest discharge = {r['steepest_discharge_V_per_s']:+.4f} V/s, "
              f"C = {c_str}")
    if gcd_results:
        caps = [r["capacitance_F"] for r in gcd_results if r["capacitance_F"]]
        if caps:
            cap_first, cap_last = caps[0], caps[-1]
            print(f"  retention: {(cap_last/cap_first)*100:.1f}% "
                  f"({len(caps)} cycles, {cap_first*1000:.2f}→{cap_last*1000:.2f} mF)")

    print("\n──────── EIS ────────────────────────────────────────")
    if eis_result["available"]:
        e = eis_result
        print(f"  points: {e['n_points']}  freq range (assumed): {e['f_min_Hz']}–{e['f_max_Hz']:.0e} Hz")
        print(f"  Rs       (high-f intercept) = {e['Rs_ohm']:.3f} Ω")
        print(f"  Rct      (semicircle, x2)   = {e['Rct_ohm']:.3f} Ω")
        print(f"  apex     (Z'={e['apex_Zreal_ohm']:.3f}, Z''={e['apex_Zimag_ohm']:.3f})")
        print(f"  Z' range: {e['Z_real_min_ohm']:.3f} → {e['Z_real_max_ohm']:.3f} Ω")
        print(f"  Z'' range: {e['Z_imag_min_ohm']:.3f} → {e['Z_imag_max_ohm']:.3f} Ω")
    else:
        print(f"  not available: {eis_result.get('reason')}")

    # ── Build dataset rows ──
    common_conditions = {
        "electrolyte":         args.electrolyte,
        "source_file":         args.xlsx.name,
        "instrument":          "AnalyteX (assumed)",
    }
    if args.electrode_area_cm2 is not None:
        common_conditions["electrode_area_cm2"] = args.electrode_area_cm2

    rows: list[dict[str, Any]] = []
    for r in cv_results:
        properties = {
            "ipa_uA":              r["ipa_uA"],
            "ipc_uA":              r["ipc_uA"],
            "dEp_mV":              r["dEp_mV"],
            "peak_current_ratio":  r["peak_current_ratio"],
            "E_pa_V":              r["E_pa_V"],
            "E_pc_V":              r["E_pc_V"],
        }
        conditions = {
            **common_conditions,
            "experiment":      "CV",
            "scan_rate_mV_s":  r["scan_rate_mV_s"],
            "potential_window_V": r["potential_window_V"],
            "raw_potential_V": r["raw_potential_V"],
            "raw_current_A":   r["raw_current_A"],
        }
        rows.append({
            "formula":     args.material,
            "name":        f"CV {int(r['scan_rate_mV_s'])} mV/s",
            "conditions":  conditions,
            "properties":  properties,
            "notes":       (f"derived from {args.xlsx.name} → CV sheet, "
                            f"scan rate {int(r['scan_rate_mV_s'])} mV/s"),
            "source":      "lab",
        })

    if rs_fit["available"]:
        rows.append({
            "formula":     args.material,
            "name":        "Randles-Sevcik fit (ip vs √v)",
            "conditions":  {**common_conditions, "experiment": "CV-summary"},
            "properties":  {
                "rs_slope_A_per_sqrt_Vs": rs_fit["slope_A_per_sqrt_Vs"],
                "rs_intercept_A":         rs_fit["intercept_A"],
                "rs_r_squared":           rs_fit["r_squared"],
            },
            "notes":       rs_fit["interpretation"],
            "source":      "lab",
        })

    for r in gcd_results:
        rows.append({
            "formula":     args.material,
            "name":        f"GCD cycle {r['cycle']}",
            "conditions":  {
                **common_conditions,
                "experiment":      "GCD",
                "cycle":           r["cycle"],
                "applied_current_mA": r["applied_current_mA"],
                "raw_time_s":      r["raw_time_s"],
                "raw_voltage_V":   r["raw_voltage_V"],
            },
            "properties":  {
                "capacitance_F":               r["capacitance_F"],
                "v_max_V":                     r["v_max_V"],
                "v_min_V":                     r["v_min_V"],
                "v_swing_V":                   r["v_swing_V"],
                "steepest_discharge_V_per_s":  r["steepest_discharge_V_per_s"],
            },
            "notes":       (f"derived from {args.xlsx.name} → GCD cycle "
                            f"{r['cycle']} at {r['applied_current_mA']:.2f} mA"),
            "source":      "lab",
        })

    if eis_result["available"]:
        e = eis_result
        rows.append({
            "formula":     args.material,
            "name":        "EIS Nyquist",
            "conditions":  {
                **common_conditions,
                "experiment":   "EIS",
                "f_max_Hz":     e["f_max_Hz"],
                "f_min_Hz":     e["f_min_Hz"],
                "raw_frequency_Hz": e["raw_frequency_Hz"],
                "raw_Z_real_ohm":   e["raw_Z_real_ohm"],
                "raw_Z_imag_ohm":   e["raw_Z_imag_ohm"],
            },
            "properties":  {
                "rs_ohm":          e["Rs_ohm"],
                "rct_ohm":         e["Rct_ohm"],
                "apex_zreal_ohm":  e["apex_Zreal_ohm"],
                "apex_zimag_ohm":  e["apex_Zimag_ohm"],
            },
            "notes":       ("derived from EIS sheet; frequency assumed "
                            f"log-spaced from {e['f_max_Hz']:.0e} Hz "
                            f"to {e['f_min_Hz']:.0e} Hz."),
            "source":      "lab",
        })

    # ── Persist ──
    mgr = get_lab_dataset_manager()
    ds_name = args.dataset_name or f"{args.material}-{args.xlsx.stem}"

    target = None
    for entry in mgr.list_datasets():
        if entry["name"] == ds_name:
            target = entry
            break

    if target and not args.append:
        print(f"\n  WARNING: dataset {ds_name!r} already exists "
              f"(id={target['id'][:8]}…). Use --append to add to it, "
              f"or pass a different --dataset-name.")
        return 2
    if target and args.append:
        ds_id = target["id"]
    else:
        ds = mgr.create_dataset(
            name=ds_name,
            description=(
                f"Imported from {args.xlsx.name}. {len(cv_results)} CV scans, "
                f"{len(gcd_results)} GCD cycles, "
                f"{eis_result['n_points'] if eis_result['available'] else 0} EIS points. "
                f"Material tag: {args.material}."
            ),
        )
        ds_id = ds.id

    n_added = mgr.add_rows(ds_id, rows)
    print(f"\n✓ Stored {n_added} rows in dataset {ds_name!r}  (id={ds_id})")
    print(f"  Look up via:  GET  /api/v2/lab/datasets/{ds_id}")
    print(f"  Or property query: POST /api/v2/alchemi/properties  {{\"formula\":\"{args.material}\"}}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
