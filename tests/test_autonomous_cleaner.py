"""
Accuracy Tests — Autonomous Data Cleaner + Calibration Analyzer
================================================================
Tests every component with known ground-truth values derived from
the actual FOG lab dataset.

Run:
    py -3.12 tests/test_autonomous_cleaner.py
"""

import sys
import json
import csv
import math
import tempfile
import shutil
from pathlib import Path
from io import StringIO

import numpy as np
import openpyxl

# ── path setup ────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.backend.ml.data_collection.autonomous_data_cleaner import (
    FormatDetector,
    CHIEISParser,
    InterleavedConcParser,
    EISCleaner,
    ElectrochemCleaner,
    DataExporter,
    AutonomousDataCleaner,
    _normalise_unit,
    _parse_conc_label,
)
from src.backend.ml.data_collection.calibration_analyzer import (
    analyze_calibration,
    _parse_conc,
    _find_peak,
    _to_uA,
)

# ── helpers ───────────────────────────────────────────────────────────────
PASS = "✅ PASS"
FAIL = "❌ FAIL"
_results = []

def check(name, condition, detail=""):
    status = PASS if condition else FAIL
    _results.append((name, condition, detail))
    print(f"  {status}  {name}" + (f"  [{detail}]" if detail else ""))
    return condition

def section(title):
    print(f"\n{'='*65}")
    print(f"  {title}")
    print(f"{'='*65}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1 — UNIT NORMALISATION
# ═══════════════════════════════════════════════════════════════════════════
section("1. Unit Normalisation")

check("µm → µM",  _normalise_unit("5 µm")  == "5 µM",  _normalise_unit("5 µm"))
check("um → µM",  _normalise_unit("10 um") == "10 µM", _normalise_unit("10 um"))
check("uM → µM",  _normalise_unit("20 uM") == "20 µM", _normalise_unit("20 uM"))
check("µl → µL",  _normalise_unit("100 µl") == "100 µL", _normalise_unit("100 µl"))
check("ul → µL",  _normalise_unit("50 ul")  == "50 µL",  _normalise_unit("50 ul"))
check("buffer unchanged", _normalise_unit("buffer") == "buffer")
check("µM unchanged",     _normalise_unit("30 µM") == "30 µM")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 — CONCENTRATION PARSING
# ═══════════════════════════════════════════════════════════════════════════
section("2. Concentration Parsing")

check("buffer → 0.0",   _parse_conc("buffer") == 0.0)
check("1 µM → 1.0",     _parse_conc("1 µM")   == 1.0)
check("5 µm → 5.0",     _parse_conc("5 µm")   == 5.0)
check("10 nM → 0.01",   abs(_parse_conc("10 nM") - 0.01) < 1e-9)
check("1 mM → 1000.0",  _parse_conc("1 mM")   == 1000.0)
check("300 µL → 300.0", _parse_conc("300 µL") == 300.0)
check("None for garbage", _parse_conc("xyz") is None)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 — FORMAT DETECTION
# ═══════════════════════════════════════════════════════════════════════════
section("3. Format Detection")

def make_ws(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    return ws

# CHI EIS
eis_ws = make_ws([
    ["Apr. 16", " 2026   13:11:18", None, None, None],
    ["A.C. Impedance", None, None, None, None],
    ["File: d:\\test.csv", None, None, None, None],
    ["Instrument Model:  CHI608E", None, None, None, None],
    ["Init E (V) = 0.386", None, None, None, None],
    ["Freq/Hz", " Z'/ohm", " Z\"/ohm", " Z/ohm", " Phase/deg"],
    [None, None, None, None, None],
    [99950, 15.48, -15.61, 21.98, -45.2],
])
check("Detect chi_eis", FormatDetector.detect(eis_ws) == "chi_eis")

# Interleaved with µM labels
conc_ws = make_ws([
    ["POTENTIAL (V)", "CURRENT (A)", None, None, None, None],
    [None, "1 µM", None, "10 µM", None, "buffer"],
    [0.1, 0.5, 0.1, 0.6, 0.1, 0.4],
    [0.2, 0.6, 0.2, 0.7, 0.2, 0.5],
])
check("Detect interleaved_conc (µM)", FormatDetector.detect(conc_ws) == "interleaved_conc")

# Interleaved with buffer only
buf_ws = make_ws([
    [None, "buffer", None, None],
    [0.1, 0.5, None, None],
])
check("Detect interleaved_conc (buffer)", FormatDetector.detect(buf_ws) == "interleaved_conc")

# Simple XY
xy_ws = make_ws([
    [0.1, 1.2e-6],
    [0.2, 1.5e-6],
    [0.3, 1.8e-6],
])
check("Detect simple_xy", FormatDetector.detect(xy_ws) == "simple_xy")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4 — CHI EIS PARSER (ground truth from actual file)
# ═══════════════════════════════════════════════════════════════════════════
section("4. CHI EIS Parser — Ground Truth")

EIS_FOG = ROOT / "Lab data/fog differet data/fog differet data/EIS FOG.xlsx"
if EIS_FOG.exists():
    wb = openpyxl.load_workbook(EIS_FOG, data_only=True)
    ws = wb.active
    parser = CHIEISParser()
    parsed = parser.parse(ws)

    check("EIS FOG: 61 data rows",   len(parsed["data"]) == 61,
          str(len(parsed["data"])))
    check("EIS FOG: 5 columns",      len(parsed["columns"]) == 5)
    check("EIS FOG: date parsed",    parsed["metadata"]["date"] is not None,
          str(parsed["metadata"]["date"]))
    check("EIS FOG: init_e = 0.18",  abs(parsed["metadata"]["init_e_v"] - 0.18) < 0.01,
          str(parsed["metadata"]["init_e_v"]))
    check("EIS FOG: freq_high = 1e5",
          abs(parsed["metadata"]["freq_high_hz"] - 1e5) < 1,
          str(parsed["metadata"]["freq_high_hz"]))
    # First row: 99950 Hz, Zreal=3.505, Zimag=-5.863
    row0 = parsed["data"][0]
    check("EIS FOG: first freq = 99950", abs(row0[0] - 99950) < 1, str(row0[0]))
    check("EIS FOG: first Zreal = 3.505", abs(row0[1] - 3.505) < 0.01, str(row0[1]))
    check("EIS FOG: first Zimag = -5.863", abs(row0[2] - (-5.863)) < 0.01, str(row0[2]))
    # Last row: 1 Hz
    rowN = parsed["data"][-1]
    check("EIS FOG: last freq = 1.0", abs(rowN[0] - 1.0) < 0.01, str(rowN[0]))
else:
    print("  ⚠ EIS FOG.xlsx not found, skipping ground-truth EIS tests")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 5 — EIS CLEANER
# ═══════════════════════════════════════════════════════════════════════════
section("5. EIS Cleaner")

# Synthetic data with known issues
raw_data = [
    [1000, 10.0, -20.0, 22.4, -63.4],   # good
    [500,  15.0, -30.0, 33.5, -63.4],   # good
    [500,  15.1, -30.1, 33.6, -63.5],   # DUPLICATE freq
    [100,  -5.0, -10.0, 11.2, -63.4],   # NEGATIVE Zreal
    [10,   50.0, -80.0, 94.3, -58.0],   # good
    [1,    200.0, -300.0, 360.6, -56.3], # good
]
parsed_synth = {
    "metadata": {},
    "columns": ["freq_hz", "zreal_ohm", "zimag_ohm", "zmag_ohm", "phase_deg"],
    "data": raw_data,
}
cleaner = EISCleaner()
cleaned = cleaner.clean(parsed_synth)

check("EIS cleaner: removed duplicate freq",
      cleaned["quality_report"]["cleaned_rows"] < 6,
      f"rows={cleaned['quality_report']['cleaned_rows']}")
check("EIS cleaner: removed negative Zreal",
      all(r[1] >= 0 for r in cleaned["data_clean"]),
      "all Zreal >= 0")
check("EIS cleaner: sorted high→low freq",
      cleaned["data_clean"][0][0] > cleaned["data_clean"][-1][0],
      f"{cleaned['data_clean'][0][0]} > {cleaned['data_clean'][-1][0]}")
check("EIS cleaner: Rs computed",
      cleaned["quality_report"]["rs_ohm"] is not None)
check("EIS cleaner: Rct computed",
      cleaned["quality_report"]["rct_ohm"] is not None)

# Test with actual FOG EIS data
if EIS_FOG.exists():
    wb = openpyxl.load_workbook(EIS_FOG, data_only=True)
    parsed_fog = CHIEISParser().parse(wb.active)
    cleaned_fog = EISCleaner().clean(parsed_fog)
    rs  = cleaned_fog["quality_report"]["rs_ohm"]
    rct = cleaned_fog["quality_report"]["rct_ohm"]
    check("EIS FOG: Rs ≈ 3.5 Ω",  abs(rs  - 3.5)   < 0.5,  f"Rs={rs}")
    check("EIS FOG: Rct ≈ 106 Ω", abs(rct - 106.0) < 10.0, f"Rct={rct}")
    check("EIS FOG: 61 clean rows", cleaned_fog["quality_report"]["cleaned_rows"] == 61)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 6 — INTERLEAVED CONCENTRATION PARSER
# ═══════════════════════════════════════════════════════════════════════════
section("6. Interleaved Concentration Parser")

DPV_FOG = ROOT / "Lab data/fog differet data/fog differet data/DPV FOG.xlsx"
if DPV_FOG.exists():
    wb = openpyxl.load_workbook(DPV_FOG, data_only=True)
    parser = InterleavedConcParser()

    # Sheet3: labelled concentrations
    ws3 = wb["Sheet3"]
    parsed3 = parser.parse(ws3)
    series3 = parsed3["series"]
    check("DPV Sheet3: 10 series detected",
          len(series3) == 10, str(list(series3.keys())))
    check("DPV Sheet3: '1 µM' present",  "1 µM"  in series3)
    check("DPV Sheet3: '70 µM' present", "70 µM" in series3)
    check("DPV Sheet3: 'buffer' present","buffer" in series3)
    check("DPV Sheet3: 53 points per series",
          len(series3["10 µM"]["potential_v"]) == 53,
          str(len(series3["10 µM"]["potential_v"])))

    # Sheet1: numeric concentration headers
    ws1 = wb["Sheet1"]
    parsed1 = parser.parse(ws1)
    series1 = parsed1["series"]
    check("DPV Sheet1: ≥9 series detected",
          len(series1) >= 9, str(list(series1.keys())))
    check("DPV Sheet1: '300 µM' present", "300 µM" in series1)
else:
    print("  ⚠ DPV FOG.xlsx not found, skipping")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 7 — ELECTROCHEM CLEANER
# ═══════════════════════════════════════════════════════════════════════════
section("7. Electrochem Cleaner")

# Synthetic series with known issues
synth_series = {
    "10 µM": {
        "potential_v": [0.3, 0.4, 0.4, 0.5, float("nan"), 0.6],  # dup + NaN
        "current_a":   [0.1, 0.2, 0.2, 0.5, 0.3,          0.4],
    },
    "buffer": {
        "potential_v": [0.3, 0.4, 0.5, 0.6],
        "current_a":   [0.1, 0.1, 0.1, 0.1],
    },
}
ec_cleaner = ElectrochemCleaner()
cleaned_ec = ec_cleaner.clean({"series": synth_series, "metadata": {}})

s10 = cleaned_ec["series_clean"]["10 µM"]
check("EC cleaner: NaN removed",
      not any(math.isnan(v) for v in s10["potential_v"]))
check("EC cleaner: duplicates merged",
      len(s10["potential_v"]) < 6,
      f"points={len(s10['potential_v'])}")
check("EC cleaner: sorted by potential",
      s10["potential_v"] == sorted(s10["potential_v"]))
check("EC cleaner: peak detected",
      cleaned_ec["quality_report"]["series"]["10 µM"]["e_peak_v"] is not None)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 8 — FULL PIPELINE (end-to-end)
# ═══════════════════════════════════════════════════════════════════════════
section("8. Full Pipeline — End-to-End")

FOG_DIR = ROOT / "Lab data/fog differet data/fog differet data"
if FOG_DIR.exists():
    with tempfile.TemporaryDirectory() as tmpdir:
        cleaner = AutonomousDataCleaner(output_dir=tmpdir)
        result  = cleaner.clean_directory(FOG_DIR)

        check("Pipeline: 6 files processed",
              result["total_files"] == 6, str(result["total_files"]))
        check("Pipeline: all 6 succeeded",
              result["success"] == 6, str(result["success"]))
        check("Pipeline: summary JSON written",
              (Path(tmpdir) / "cleaning_summary.json").exists())

        # Check EIS outputs
        eis_csv = list(Path(tmpdir).rglob("*eis_fog*.csv"))
        check("Pipeline: EIS FOG CSV generated", len(eis_csv) > 0)
        if eis_csv:
            rows = list(csv.reader(eis_csv[0].read_text().splitlines()))
            check("Pipeline: EIS CSV has header",
                  rows[0] == ["freq_hz", "zreal_ohm", "zimag_ohm", "zmag_ohm", "phase_deg"])
            check("Pipeline: EIS CSV has 61 data rows",
                  len(rows) - 1 == 61, str(len(rows) - 1))

        # Check DPV outputs
        dpv_csvs = list(Path(tmpdir).rglob("*Sheet3*10_µM*.csv"))
        check("Pipeline: DPV 10µM CSV generated", len(dpv_csvs) > 0)
        if dpv_csvs:
            rows = list(csv.reader(dpv_csvs[0].read_text().splitlines()))
            check("Pipeline: DPV CSV has header",
                  rows[0] == ["potential_v", "current_a"])
            check("Pipeline: DPV CSV has 53 data rows",
                  len(rows) - 1 == 53, str(len(rows) - 1))

        # Check Gomutra outputs
        gom_csvs = list(Path(tmpdir).rglob("*300_µL*.csv"))
        check("Pipeline: Gomutra 300µL CSV generated", len(gom_csvs) > 0)
else:
    print("  ⚠ FOG directory not found, skipping end-to-end test")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 9 — CALIBRATION ANALYZER
# ═══════════════════════════════════════════════════════════════════════════
section("9. Calibration Analyzer — Accuracy")

# Test _find_peak
pot = np.array([0.3, 0.4, 0.5, 0.6, 0.7])
cur = np.array([0.1, 0.2, 0.8, 0.3, 0.1])
ep, ip = _find_peak(pot, cur)
check("find_peak: correct E_peak", abs(ep - 0.5) < 1e-9, str(ep))
check("find_peak: correct I_peak", abs(ip - 0.8) < 1e-9, str(ip))

# Test _to_uA
check("to_uA: A→µA conversion",
      abs(_to_uA(np.array([1e-6]))[0] - 1.0) < 1e-9)
check("to_uA: µA unchanged",
      abs(_to_uA(np.array([1.0]))[0] - 1.0) < 1e-9)

# Test _parse_conc
check("parse_conc: 10 µM", _parse_conc("10 µM") == 10.0)
check("parse_conc: buffer", _parse_conc("buffer") == 0.0)
check("parse_conc: 1 nM",   abs(_parse_conc("1 nM") - 0.001) < 1e-9)

# Test with actual cleaned DPV data
CLEANED_DPV = ROOT / "data/cleaned/fog/DPV FOG/DPV FOG_Sheet3_all.json"
if CLEANED_DPV.exists():
    result = analyze_calibration(CLEANED_DPV)
    check("Calibration: no error",       "error" not in result)
    check("Calibration: R² > 0.99",      result.get("r_squared", 0) > 0.99,
          f"R²={result.get('r_squared')}")
    check("Calibration: sensitivity > 0",
          result.get("sensitivity_uA_per_uM", 0) > 0,
          f"sens={result.get('sensitivity_uA_per_uM')}")
    check("Calibration: LOD computed",   result.get("lod_uM") is not None)
    check("Calibration: LOQ > LOD",
          (result.get("loq_uM") or 0) > (result.get("lod_uM") or 0))
    check("Calibration: 9 non-buffer points",
          result.get("n_points", 0) == 9, str(result.get("n_points")))
    # Peak at ~0.48 V for 10 µM
    pt = result["peak_table"].get("10 µM", {})
    check("Calibration: 10µM E_peak ≈ 0.48 V",
          abs(pt.get("e_peak_v", 0) - 0.48) < 0.02,
          f"E_peak={pt.get('e_peak_v')}")
else:
    print("  ⚠ Cleaned DPV JSON not found — run cleaner first")

# Test with Gomutra data
CLEANED_GOM = ROOT / "data/cleaned/fog/GOMUTRA CONCENTRATION STUDIES/GOMUTRA CONCENTRATION STUDIES_Sheet1_all.json"
if CLEANED_GOM.exists():
    result_g = analyze_calibration(CLEANED_GOM)
    check("Gomutra calibration: R² > 0.99",
          result_g.get("r_squared", 0) > 0.99,
          f"R²={result_g.get('r_squared')}")
    check("Gomutra calibration: 11 non-buffer points",
          result_g.get("n_points", 0) == 11, str(result_g.get("n_points")))


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 10 — DATA INTEGRITY CHECKS
# ═══════════════════════════════════════════════════════════════════════════
section("10. Data Integrity — Cleaned Files")

CLEAN_DIR = ROOT / "data/cleaned/fog"
if CLEAN_DIR.exists():
    # EIS files: check all 4 have correct Rs/Rct ordering
    eis_meta_files = list(CLEAN_DIR.rglob("*_meta.json"))
    check("Integrity: 4 EIS meta files", len(eis_meta_files) == 4,
          str(len(eis_meta_files)))

    rct_values = {}
    for mf in eis_meta_files:
        meta = json.loads(mf.read_text())
        qr   = meta.get("quality_report", {})
        name = mf.stem.replace("_meta", "")
        rct_values[name] = qr.get("rct_ohm", 0)

    # FOG should have lowest Rct
    if rct_values:
        fog_key = min(rct_values, key=lambda k: rct_values[k])
        check("Integrity: FOG has lowest Rct",
              "fog" in fog_key.lower() or "eis_fog" in fog_key.lower(),
              f"lowest={fog_key} ({rct_values[fog_key]:.1f}Ω)")

    # DPV: check all 10 concentration CSVs exist
    dpv_csvs = list((CLEAN_DIR / "DPV FOG").glob("*Sheet3*.csv"))
    check("Integrity: 10 DPV Sheet3 CSVs",
          len(dpv_csvs) == 10, str(len(dpv_csvs)))

    # Gomutra: check all 12 concentration CSVs exist
    gom_csvs = list((CLEAN_DIR / "GOMUTRA CONCENTRATION STUDIES").glob("*.csv"))
    check("Integrity: 12 Gomutra CSVs",
          len(gom_csvs) == 12, str(len(gom_csvs)))

    # Check no NaN in any CSV
    nan_found = []
    for csv_file in CLEAN_DIR.rglob("*.csv"):
        content = csv_file.read_text(encoding="utf-8")
        if "nan" in content.lower() or "inf" in content.lower():
            nan_found.append(csv_file.name)
    check("Integrity: no NaN/Inf in any CSV",
          len(nan_found) == 0, str(nan_found) if nan_found else "clean")

    # Check EIS CSV values are physically reasonable
    eis_fog_csv = list(CLEAN_DIR.rglob("*eis_fog*.csv"))
    if eis_fog_csv:
        rows = list(csv.reader(eis_fog_csv[0].read_text().splitlines()))
        freqs  = [float(r[0]) for r in rows[1:]]
        zreals = [float(r[1]) for r in rows[1:]]
        zimajs = [float(r[2]) for r in rows[1:]]
        check("Integrity: EIS freqs all positive",  all(f > 0 for f in freqs))
        check("Integrity: EIS Zreal all positive",  all(z > 0 for z in zreals))
        check("Integrity: EIS Zimag all negative",  all(z < 0 for z in zimajs))
        check("Integrity: EIS freq sorted desc",
              freqs == sorted(freqs, reverse=True))
else:
    print("  ⚠ Cleaned data directory not found — run cleaner first")


# ═══════════════════════════════════════════════════════════════════════════
# FINAL REPORT
# ═══════════════════════════════════════════════════════════════════════════
total   = len(_results)
passed  = sum(1 for _, ok, _ in _results if ok)
failed  = total - passed
pct     = 100 * passed / total if total else 0

print(f"\n{'='*65}")
print(f"  FINAL RESULTS")
print(f"{'='*65}")
print(f"  Passed:  {passed}/{total}  ({pct:.1f}%)")
print(f"  Failed:  {failed}")

if failed:
    print(f"\n  Failed tests:")
    for name, ok, detail in _results:
        if not ok:
            print(f"    ✗ {name}" + (f"  [{detail}]" if detail else ""))

print()

if __name__ == "__main__":
    sys.exit(0 if failed == 0 else 1)
